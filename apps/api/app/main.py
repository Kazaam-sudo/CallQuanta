import csv
import hashlib
import hmac
import json
import secrets
import logging
import mimetypes
import os
import re
import time
from pathlib import Path
from urllib.parse import quote
from uuid import uuid4
from datetime import UTC, datetime, timedelta, time as datetime_time

import redis
import requests
import yaml
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from fastapi.responses import JSONResponse, StreamingResponse
from io import BytesIO, StringIO
from openpyxl import Workbook
from sqlalchemy import create_engine, func, or_, select, text, delete
from sqlalchemy.orm import Session, sessionmaker
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from .db import AppSetting, AuditEvent, Base, Call, IngestionEvent, ProviderConfig, QACoachingAction, QAFeedback, QAReview, QAReviewAssignment, ScorecardConfig, SttProviderConfig, TelephonyIntegration, TranscriptSegment, User, migrate_access_control_tables, migrate_calls_table, migrate_production_readiness_tables, migrate_qa_coaching_actions_table, migrate_pilot_feedback_tables, migrate_qa_reviews_table, migrate_stt_provider_configs_table, migrate_telephony_ingestion_tables
from .stt_languages import SUPPORTED_STT_LANGUAGES, SUPPORTED_STT_LANGUAGE_CODES, normalize_language_code, normalize_stt_language

app = FastAPI(title="CallQuanta API", version="0.24.3")
logger = logging.getLogger("callquanta.api")

DATABASE_URL = "postgresql+psycopg://callquanta:callquanta@postgres:5432/callquanta"
DATABASE_URL = os.environ.get("DATABASE_URL", DATABASE_URL)
REDIS_URL = os.environ.get("REDIS_URL", "redis://redis:6379/0")
TRANSCRIPTION_QUEUE = os.environ.get("TRANSCRIPTION_QUEUE", "transcription_jobs")
RECORDING_QUEUE = os.environ.get("RECORDING_QUEUE", "recording_jobs")
QA_QUEUE = os.environ.get("QA_QUEUE", "qa_jobs")
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "uploads"))
CORS_ORIGINS = os.environ.get(
    "CORS_ORIGINS",
    "http://localhost:3000,http://127.0.0.1:3000",
)
ALLOWED_CORS_ORIGINS = [origin.strip() for origin in CORS_ORIGINS.split(",") if origin.strip()]
ALLOWED_UPLOAD_EXTENSIONS = {".wav", ".mp3", ".m4a", ".ogg", ".opus", ".flac", ".webm"}
MAX_DISPLAY_FILENAME_LENGTH = 255
INGESTION_PAYLOAD_LOGGING = os.environ.get("INGESTION_PAYLOAD_LOGGING", "false").lower() in {"1", "true", "yes", "on"}
ALLOWED_UPLOAD_CONTENT_TYPES = {
    "audio/wav",
    "audio/x-wav",
    "audio/mpeg",
    "audio/mp3",
    "audio/mp4",
    "audio/aac",
    "audio/ogg",
    "application/ogg",
    "audio/opus",
    "audio/flac",
    "audio/webm",
    "video/webm",
}
UPLOAD_CHUNK_SIZE = 1024 * 1024
AUDIO_STREAM_CHUNK_SIZE = 64 * 1024
DEFAULT_MAX_UPLOAD_BYTES_PER_FILE = 100 * 1024 * 1024
DEFAULT_MAX_BULK_UPLOAD_BYTES = 500 * 1024 * 1024
MAX_UPLOAD_BYTES_PER_FILE = int(
    os.environ.get(
        "MAX_UPLOAD_BYTES_PER_FILE",
        os.environ.get("MAX_UPLOAD_BYTES", str(DEFAULT_MAX_UPLOAD_BYTES_PER_FILE)),
    )
    or "0"
)
MAX_BULK_UPLOAD_BYTES = int(os.environ.get("MAX_BULK_UPLOAD_BYTES", str(DEFAULT_MAX_BULK_UPLOAD_BYTES)) or "0")

APP_ENV = os.environ.get("APP_ENV", "development").lower()
REQUIRE_AUTH = os.environ.get("REQUIRE_AUTH", "true").lower() in {"1", "true", "yes", "on"}
SESSION_SECRET = os.environ.get("SESSION_SECRET", "dev-session-secret-change-me")
SESSION_COOKIE_NAME = os.environ.get("SESSION_COOKIE_NAME", "callquanta_session")
SESSION_TTL_SECONDS = int(os.environ.get("SESSION_TTL_SECONDS", str(7 * 24 * 60 * 60)))
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "").strip().lower()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
WORKER_HEARTBEAT_SECONDS = int(os.environ.get("WORKER_HEARTBEAT_SECONDS", "30"))
RETENTION_SETTINGS_KEY = "retention"
DEFAULT_RETENTION_SETTINGS = {
    "audio_days": None,
    "transcripts_days": None,
    "qa_reviews_days": None,
    "ingestion_events_days": None,
}

LANGUAGE_CATALOG = [
    {"code": "en", "label": "English", "native_label": "English", "ui_supported": True, "llm_supported": True},
    {"code": "ru", "label": "Russian", "native_label": "Русский", "ui_supported": True, "llm_supported": True},
    {"code": "uz", "label": "Uzbek", "native_label": "O‘zbek", "ui_supported": True, "llm_supported": True},
    {"code": "es", "label": "Spanish", "native_label": "Español", "ui_supported": False, "llm_supported": True},
    {"code": "pt", "label": "Portuguese", "native_label": "Português", "ui_supported": False, "llm_supported": True},
    {"code": "de", "label": "German", "native_label": "Deutsch", "ui_supported": False, "llm_supported": True},
    {"code": "fr", "label": "French", "native_label": "Français", "ui_supported": False, "llm_supported": True},
    {"code": "tr", "label": "Turkish", "native_label": "Türkçe", "ui_supported": False, "llm_supported": True},
    {"code": "ar", "label": "Arabic", "native_label": "العربية", "ui_supported": False, "llm_supported": True},
    {"code": "custom", "label": "Custom", "native_label": "Custom", "ui_supported": False, "llm_supported": True},
]
LANGUAGE_LABEL_BY_CODE = {item["code"]: item["label"] for item in LANGUAGE_CATALOG}
WORKSPACE_SETTINGS_KEY = "workspace"
DEFAULT_WORKSPACE_SETTINGS = {
    "interface_language": "en",
    "qa_report_language_mode": "workspace",
    "qa_report_language": "English",
}


TELEPHONY_PROVIDER_TYPES = {"generic_webhook", "voximplant", "asterisk", "twilio", "zadarma", "custom"}
TELEPHONY_PROVIDER_PRESETS = [
    {"id": "generic_webhook", "label": "Generic webhook", "implemented": True},
    {"id": "voximplant", "label": "Voximplant", "implemented": False},
    {"id": "asterisk", "label": "Asterisk", "implemented": False},
    {"id": "twilio", "label": "Twilio", "implemented": False},
    {"id": "zadarma", "label": "Zadarma", "implemented": False},
    {"id": "custom", "label": "Custom", "implemented": False},
]

STT_PROVIDER_TYPES = {
    "faster_whisper_local",
    "openai_compatible_audio",
    "groq_whisper",
    "deepgram",
    "assemblyai",
    "google_speech",
    "azure_speech",
    "custom",
}
STT_PROVIDER_PRESETS = [
    {"id": "local_faster_whisper", "label": "Local faster-whisper", "provider_type": "faster_whisper_local", "default_base_url": "", "default_model": "tiny", "api_key_required": False, "note": "Runs inside your STT worker using local faster-whisper settings."},
    {"id": "openai_audio", "label": "OpenAI audio transcription", "provider_type": "openai_compatible_audio", "default_base_url": "https://api.openai.com/v1", "default_model": "whisper-1", "api_key_required": True},
    {"id": "groq_whisper", "label": "Groq Whisper", "provider_type": "groq_whisper", "default_base_url": "https://api.groq.com/openai/v1", "default_model": "", "api_key_required": True, "note": "Preset can be saved now; transcription support is not implemented yet."},
    {"id": "deepgram", "label": "Deepgram", "provider_type": "deepgram", "default_base_url": "https://api.deepgram.com/v1", "default_model": "", "api_key_required": True, "note": "Preset can be saved now; transcription support is not implemented yet."},
    {"id": "assemblyai", "label": "AssemblyAI", "provider_type": "assemblyai", "default_base_url": "https://api.assemblyai.com/v2", "default_model": "", "api_key_required": True, "note": "Preset can be saved now; transcription support is not implemented yet."},
    {"id": "google_speech", "label": "Google Speech-to-Text", "provider_type": "google_speech", "default_base_url": "", "default_model": "", "api_key_required": True, "note": "Preset can be saved now; transcription support is not implemented yet."},
    {"id": "azure_speech", "label": "Azure Speech", "provider_type": "azure_speech", "default_base_url": "", "default_model": "", "api_key_required": True, "note": "Preset can be saved now; transcription support is not implemented yet."},
    {"id": "custom", "label": "Custom STT provider", "provider_type": "custom", "default_base_url": "", "default_model": "", "api_key_required": False, "note": "Placeholder for future custom STT integrations."},
]

PROVIDER_PRESETS = [
    {"id": "ollama", "label": "Ollama Local", "provider_type": "openai_compatible", "default_base_url": "http://ollama:11434/v1", "default_model": "qwen2.5:1.5b", "api_key_required": False},
    {"id": "openai", "label": "OpenAI", "provider_type": "openai_compatible", "default_base_url": "https://api.openai.com/v1", "default_model": "gpt-5.4-nano", "api_key_required": True},
    {"id": "groq", "label": "Groq", "provider_type": "openai_compatible", "default_base_url": "https://api.groq.com/openai/v1", "default_model": "qwen/qwen3-32b", "api_key_required": True},
    {"id": "openrouter", "label": "OpenRouter", "provider_type": "openai_compatible", "default_base_url": "https://openrouter.ai/api/v1", "default_model": "openai/gpt-oss-120b", "api_key_required": True},
    {"id": "cloudflare", "label": "Cloudflare Workers AI", "provider_type": "openai_compatible", "default_base_url": "https://api.cloudflare.com/client/v4/accounts/{ACCOUNT_ID}/ai/v1", "default_model": "@cf/meta/llama-3.1-8b-instruct", "api_key_required": True, "note": "Replace {ACCOUNT_ID} before using."},
    {"id": "gemini", "label": "Google Gemini", "provider_type": "openai_compatible", "default_base_url": "https://generativelanguage.googleapis.com/v1beta/openai", "default_model": "gemini-3.5-flash", "api_key_required": True},
    {"id": "anthropic", "label": "Anthropic Claude", "provider_type": "openai_compatible", "default_base_url": "https://api.anthropic.com/v1", "default_model": "claude-sonnet-4-5", "api_key_required": True, "note": "Compatibility may not support every OpenAI parameter."},
    {"id": "together", "label": "Together AI", "provider_type": "openai_compatible", "default_base_url": "https://api.together.xyz/v1", "default_model": "Qwen/Qwen3-32B", "api_key_required": True},
    {"id": "custom", "label": "Custom OpenAI-compatible", "provider_type": "openai_compatible", "default_base_url": "", "default_model": "", "api_key_required": False},
]


app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
redis_client = redis.Redis.from_url(REDIS_URL, decode_responses=True)


class ProviderUpsertRequest(BaseModel):
    id: int | None = None
    name: str
    preset: str
    provider_type: str = "openai_compatible"
    base_url: str
    model: str
    api_key: str | None = None
    timeout_seconds: float = 180
    is_active: bool = False


class ProviderTestRequest(BaseModel):
    provider_type: str = "openai_compatible"
    base_url: str
    model: str
    api_key: str | None = None
    timeout_seconds: float = 60


class SttProviderUpsertRequest(BaseModel):
    id: int | None = None
    name: str
    provider_type: str = "faster_whisper_local"
    preset: str = "local_faster_whisper"
    model: str = "tiny"
    base_url: str | None = None
    api_key: str | None = None
    timeout_seconds: int = 180
    is_active: bool = False


class SttProviderTestRequest(BaseModel):
    id: int | None = None
    name: str | None = None
    provider_type: str = "faster_whisper_local"
    preset: str = "local_faster_whisper"
    model: str = "tiny"
    base_url: str | None = None
    api_key: str | None = None
    timeout_seconds: int = 60


DEFAULT_SCORECARD_PATH = Path(os.environ.get("SCORECARD_PATH", "/app/packages/scorecards/default_sales_qa.yaml"))


class ScorecardCriterion(BaseModel):
    id: str
    title: str
    max_points: float
    description: str
    positive_examples: list[str] | None = None
    negative_examples: list[str] | None = None


class ScorecardPayload(BaseModel):
    name: str
    report_language: str = "english"
    criteria: list[ScorecardCriterion]




class WorkspaceSettingsPayload(BaseModel):
    interface_language: str | None = None
    qa_report_language_mode: str | None = None
    qa_report_language: str | None = None

class CallMetadataPayload(BaseModel):
    agent_name: str | None = None
    team: str | None = None
    campaign: str | None = None
    direction: str | None = None
    language: str | None = None




class TelephonyIntegrationPayload(BaseModel):
    id: int | None = None
    name: str
    provider_type: str = "generic_webhook"
    is_active: bool = True
    auto_transcribe: bool = True
    auto_analyze: bool = False
    default_agent_name: str | None = None
    default_team: str | None = None
    default_campaign: str | None = None
    default_direction: str | None = None
    default_language: str | None = None
    generate_token: bool = False


class TelephonyImportPayload(BaseModel):
    external_call_id: str
    recording_url: str
    filename: str | None = None
    agent_name: str | None = None
    team: str | None = None
    campaign: str | None = None
    direction: str | None = None
    language: str | None = None
    customer_phone: str | None = None
    agent_phone: str | None = None
    started_at: datetime | None = None
    ended_at: datetime | None = None
    duration_seconds: int | None = None
    auto_transcribe: bool | None = None
    auto_analyze: bool | None = None


class BatchCallsPayload(BaseModel):
    call_ids: list[int]


class BatchMetadataPayload(CallMetadataPayload):
    call_ids: list[int]


class BatchDeletePayload(BaseModel):
    call_ids: list[int]
    delete_files: bool = True


VALID_ROLES = {"admin", "manager", "supervisor", "agent", "viewer"}
VALID_VISIBILITY_SCOPES = {"all", "team", "own"}
ROLE_DEFAULT_SCOPE = {"admin": "all", "manager": "team", "supervisor": "team", "agent": "own", "viewer": "team"}


class UserAccessCreatePayload(BaseModel):
    email: str
    password: str | None = None
    display_name: str | None = None
    role: str = "viewer"
    team: str | None = None
    agent_name: str | None = None
    visibility_scope: str | None = None
    is_active: bool = True
    must_change_password: bool = False


class UserAccessPatchPayload(BaseModel):
    email: str | None = None
    display_name: str | None = None
    role: str | None = None
    team: str | None = None
    agent_name: str | None = None
    visibility_scope: str | None = None
    is_active: bool | None = None
    must_change_password: bool | None = None


class ResetPasswordPayload(BaseModel):
    temporary_password: str | None = None


class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str


def get_db() -> Session:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class LoginPayload(BaseModel):
    email: str
    password: str


class RetentionSettingsPayload(BaseModel):
    audio_days: int | None = None
    transcripts_days: int | None = None
    qa_reviews_days: int | None = None
    ingestion_events_days: int | None = None


class CriterionHumanReviewPayload(BaseModel):
    criterion_id: str | None = None
    criterion_index: int | None = None
    human_score: float | None = None
    human_comment: str | None = None
    human_severity: str | None = None
    human_agrees: bool | None = None


class HumanReviewPayload(BaseModel):
    review_status: str = "human_reviewed"
    human_total_score: float | None = None
    human_summary: str | None = None
    human_notes: str | None = None
    calibration_flag: bool | None = None
    calibration_notes: str | None = None
    criteria: list[CriterionHumanReviewPayload] | None = None


class CoachingActionPayload(BaseModel):
    title: str
    description: str | None = None
    assigned_to_user_id: int | None = None
    due_date: datetime | None = None


class CoachingActionPatchPayload(BaseModel):
    title: str | None = None
    description: str | None = None
    assigned_to_user_id: int | None = None
    due_date: datetime | None = None
    status: str | None = None


class QAReviewAssignmentPayload(BaseModel):
    assigned_to_user_id: int
    due_date: datetime | None = None


class QAReviewAssignmentPatchPayload(BaseModel):
    status: str | None = None
    due_date: datetime | None = None


class QAFeedbackPayload(BaseModel):
    transcript_quality: str | None = None
    qa_analysis_quality: str | None = None
    score_agreement: str | None = None
    scorecard_fit: str | None = None
    ai_missed_something: bool | None = None
    ai_missed_comment: str | None = None
    ai_false_positive: bool | None = None
    ai_false_positive_comment: str | None = None
    useful_for_coaching: bool | None = None
    coaching_usefulness_comment: str | None = None
    overall_feedback: str | None = None
    issue_tags: list[str] | None = None



AUTH_EXEMPT_PATHS = {"/health", "/health/ready", "/auth/login", "/auth/logout"}
AUTH_EXEMPT_PREFIXES = ("/integrations/telephony/webhook/",)


def _utcnow() -> datetime:
    return datetime.now(UTC)


def _normalize_email(email: str | None) -> str:
    return (email or "").strip().lower()


def _validate_email(email: str | None) -> str:
    normalized = _normalize_email(email)
    if not normalized:
        raise HTTPException(status_code=400, detail="email is required")
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", normalized):
        raise HTTPException(status_code=400, detail="email must be a valid email address")
    return normalized


def _validate_password_policy(password: str | None, field_name: str = "password") -> str:
    value = password or ""
    if len(value) < 8:
        raise HTTPException(status_code=400, detail=f"{field_name} must be at least 8 characters")
    if not re.search(r"[A-Za-zА-Яа-я]", value) or not re.search(r"\d", value):
        raise HTTPException(status_code=400, detail=f"{field_name} must contain letters and numbers")
    return value


def _hash_password(password: str) -> str:
    return generate_password_hash(password, method="pbkdf2:sha256", salt_length=16)


def _verify_password(password: str, password_hash: str) -> bool:
    return check_password_hash(password_hash, password)


def _session_signature(payload: str) -> str:
    return hmac.new(SESSION_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def _create_session_token(user: User) -> str:
    expires_at = int(time.time()) + SESSION_TTL_SECONDS
    payload = f"{user.id}:{expires_at}:{secrets.token_urlsafe(8)}"
    return f"{payload}.{_session_signature(payload)}"


def _decode_session_token(token: str | None) -> int | None:
    if not token or "." not in token:
        return None
    payload, signature = token.rsplit(".", 1)
    if not hmac.compare_digest(_session_signature(payload), signature):
        return None
    parts = payload.split(":", 2)
    if len(parts) < 2:
        return None
    try:
        user_id = int(parts[0])
        expires_at = int(parts[1])
    except ValueError:
        return None
    if expires_at < int(time.time()):
        return None
    return user_id


def _get_user_from_request(request: Request, db: Session) -> User | None:
    user_id = _decode_session_token(request.cookies.get(SESSION_COOKIE_NAME))
    if not user_id:
        return None
    user = db.get(User, user_id)
    if not user or not user.is_active:
        return None
    return user


def get_current_user(request: Request, db: Session = Depends(get_db)) -> User:
    if not REQUIRE_AUTH:
        user = db.execute(select(User).where(User.is_active == True).limit(1)).scalar_one_or_none()
        if user:
            return user
        return User(id=0, email="development@callquanta.local", role="admin", is_active=True, password_hash="")
    user = _get_user_from_request(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return user


def _user_scope(user: User) -> str:
    if user.role == "admin":
        return "all"
    scope = (getattr(user, "visibility_scope", None) or ROLE_DEFAULT_SCOPE.get(user.role, "team")).strip().lower()
    return scope if scope in VALID_VISIBILITY_SCOPES else ROLE_DEFAULT_SCOPE.get(user.role, "team")


def _normalize_role(role: str | None) -> str:
    normalized = (role or "viewer").strip().lower()
    if normalized not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"role must be one of: {', '.join(sorted(VALID_ROLES))}")
    return normalized


def _normalize_visibility_scope(scope: str | None, role: str) -> str:
    if role == "admin":
        return "all"
    normalized = (scope or ROLE_DEFAULT_SCOPE.get(role, "team")).strip().lower()
    if normalized not in VALID_VISIBILITY_SCOPES:
        raise HTTPException(status_code=400, detail="visibility_scope must be one of: all, team, own")
    return normalized


def _can_view_call(user: User, call: Call) -> bool:
    scope = _user_scope(user)
    if user.role == "admin" or scope == "all":
        return True
    if scope == "team":
        team = _normalize_optional_text(getattr(user, "team", None))
        return bool(team) and (call.team or "").strip() == team
    if scope == "own":
        agent_name = _normalize_optional_text(getattr(user, "agent_name", None))
        return bool(agent_name) and (call.agent_name or "").strip() == agent_name
    return False


def apply_call_scope(stmt, user: User):
    scope = _user_scope(user)
    if user.role == "admin" or scope == "all":
        return stmt
    if scope == "team":
        team = _normalize_optional_text(getattr(user, "team", None))
        return stmt.where(Call.team == team) if team else stmt.where(Call.id < 0)
    if scope == "own":
        agent_name = _normalize_optional_text(getattr(user, "agent_name", None))
        return stmt.where(Call.agent_name == agent_name) if agent_name else stmt.where(Call.id < 0)
    return stmt.where(Call.id < 0)


def require_can_view_call(call: Call, user: User) -> Call:
    if not _can_view_call(user, call):
        raise HTTPException(status_code=403, detail="Call is outside your visibility scope")
    return call


def require_can_modify_call(call: Call, user: User) -> Call:
    if user.role in {"viewer", "agent"}:
        raise HTTPException(status_code=403, detail="You do not have permission to modify calls")
    return require_can_view_call(call, user)


def _serialize_user(user: User) -> dict:
    return {
        "id": user.id,
        "email": user.email,
        "display_name": getattr(user, "display_name", None),
        "role": user.role,
        "team": getattr(user, "team", None),
        "agent_name": getattr(user, "agent_name", None),
        "visibility_scope": _user_scope(user),
        "is_active": user.is_active,
        "must_change_password": bool(getattr(user, "must_change_password", False)),
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "updated_at": user.updated_at.isoformat() if user.updated_at else None,
    }


def _log_audit_event(db: Session, actor: User | None, action: str, entity_type: str, entity_id: object | None = None, message: str | None = None, metadata: dict | None = None, commit: bool = False) -> None:
    try:
        db.add(AuditEvent(
            actor_user_id=getattr(actor, "id", None),
            actor_email=getattr(actor, "email", None),
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            message=message,
            metadata_json=metadata,
        ))
        if commit:
            db.commit()
    except Exception:
        logger.exception("Failed to write audit event %s", action)


def _generate_temporary_password() -> str:
    alphabet = "abcdefghijkmnopqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return f"Cq{secrets.randbelow(90) + 10}" + "".join(secrets.choice(alphabet) for _ in range(10))


def _auth_is_exempt(path: str) -> bool:
    return path in AUTH_EXEMPT_PATHS or any(path.startswith(prefix) for prefix in AUTH_EXEMPT_PREFIXES)


@app.middleware("http")
async def require_auth_middleware(request: Request, call_next):
    if request.method == "OPTIONS" or not REQUIRE_AUTH or _auth_is_exempt(request.url.path):
        return await call_next(request)
    with SessionLocal() as db:
        user = _get_user_from_request(request, db)
        if not user:
            return JSONResponse({"detail": "Authentication required"}, status_code=401)
        if getattr(user, "must_change_password", False) and request.url.path not in {"/auth/change-password", "/auth/logout", "/auth/me"}:
            return JSONResponse({"detail": "Password change required", "must_change_password": True}, status_code=403)
        request.state.user = _serialize_user(user)
    return await call_next(request)


def _bootstrap_admin_user(db: Session) -> None:
    has_user = db.execute(select(User.id).limit(1)).first()
    if has_user:
        return
    if not ADMIN_EMAIL or not ADMIN_PASSWORD:
        message = "No users exist and ADMIN_EMAIL/ADMIN_PASSWORD are not fully configured; login will be unavailable."
        if APP_ENV in {"production", "prod"}:
            logger.warning("%s Set ADMIN_EMAIL and ADMIN_PASSWORD before production use.", message)
        else:
            logger.warning("%s Development mode allows you to set them and restart.", message)
        return
    db.add(User(email=ADMIN_EMAIL, password_hash=_hash_password(ADMIN_PASSWORD), role="admin", visibility_scope="all", is_active=True, must_change_password=False))
    db.commit()
    logger.info("Created first admin user from ADMIN_EMAIL: %s", ADMIN_EMAIL)


def _normalize_workspace_settings(value: dict | None) -> dict:
    settings = dict(DEFAULT_WORKSPACE_SETTINGS)
    if isinstance(value, dict):
        settings.update({k: v for k, v in value.items() if v is not None})
    if settings.get("interface_language") not in LANGUAGE_LABEL_BY_CODE:
        settings["interface_language"] = "en"
    if settings.get("qa_report_language_mode") not in {"workspace", "same_as_transcript", "custom"}:
        settings["qa_report_language_mode"] = "workspace"
    language = str(settings.get("qa_report_language") or "").strip()
    if not language:
        language = LANGUAGE_LABEL_BY_CODE.get(settings["interface_language"], "English")
    settings["qa_report_language"] = language
    return settings


def _get_workspace_settings(db: Session) -> dict:
    setting = db.get(AppSetting, WORKSPACE_SETTINGS_KEY)
    return _normalize_workspace_settings(setting.value if setting else None)


def _upsert_workspace_settings(db: Session, payload: WorkspaceSettingsPayload) -> dict:
    current = _get_workspace_settings(db)
    updates = payload.model_dump(exclude_unset=True)
    if "interface_language" in updates:
        code = str(updates["interface_language"] or "").strip()
        if code not in LANGUAGE_LABEL_BY_CODE:
            raise HTTPException(status_code=400, detail="Unsupported interface_language")
        current["interface_language"] = code
        if current.get("qa_report_language_mode") == "workspace" and "qa_report_language" not in updates:
            current["qa_report_language"] = LANGUAGE_LABEL_BY_CODE.get(code, "English")
    if "qa_report_language_mode" in updates:
        mode = str(updates["qa_report_language_mode"] or "").strip()
        if mode not in {"workspace", "same_as_transcript", "custom"}:
            raise HTTPException(status_code=400, detail="qa_report_language_mode must be one of: workspace, same_as_transcript, custom")
        current["qa_report_language_mode"] = mode
    if "qa_report_language" in updates:
        language = str(updates["qa_report_language"] or "").strip()
        if not language:
            raise HTTPException(status_code=400, detail="qa_report_language must not be empty")
        current["qa_report_language"] = language
    current = _normalize_workspace_settings(current)
    setting = db.get(AppSetting, WORKSPACE_SETTINGS_KEY)
    if not setting:
        setting = AppSetting(key=WORKSPACE_SETTINGS_KEY, value=current)
        db.add(setting)
    else:
        setting.value = current
    db.commit()
    return current

def _provider_test_request(provider_type: str, base_url: str, model: str, api_key: str | None, timeout_seconds: float) -> dict:
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    started = time.perf_counter()
    try:
        response = requests.post(
            f"{base_url.rstrip('/')}/chat/completions",
            headers=headers,
            json={"model": model, "temperature": 0, "messages": [{"role": "user", "content": "Return only JSON: {\"ok\": true}"}]},
            timeout=timeout_seconds,
        )
        latency_ms = int((time.perf_counter() - started) * 1000)
        if response.ok:
            return {"ok": True, "latency_ms": latency_ms, "model": model}

        provider_error: object
        try:
            provider_error = response.json()
        except ValueError:
            provider_error = (response.text or "")[:1000]
        return {
            "ok": False,
            "status_code": response.status_code,
            "latency_ms": latency_ms,
            "model": model,
            "provider_error": provider_error,
        }
    except requests.RequestException as exc:
        latency_ms = int((time.perf_counter() - started) * 1000)
        return {"ok": False, "latency_ms": latency_ms, "model": model, "provider_error": str(exc)[:300]}


@app.get("/health")
def health() -> dict[str, str]:
    logger.info("Health check requested")
    return {"status": "ok"}




def _check_db_ready() -> tuple[bool, str | None]:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, None
    except Exception as exc:
        return False, str(exc)[:300]


def _check_redis_ready() -> tuple[bool, str | None]:
    try:
        redis_client.ping()
        return True, None
    except redis.RedisError as exc:
        return False, str(exc)[:300]


@app.get("/health/ready")
def health_ready(response: Response) -> dict:
    db_ok, db_error = _check_db_ready()
    redis_ok, redis_error = _check_redis_ready()
    ready = db_ok and redis_ok
    if not ready:
        response.status_code = 503
    return {"status": "ready" if ready else "not_ready", "checks": {"db": {"ok": db_ok, "error": db_error}, "redis": {"ok": redis_ok, "error": redis_error}}}


@app.post("/auth/login")
def login(payload: LoginPayload, response: Response, db: Session = Depends(get_db)) -> dict:
    email = _normalize_email(payload.email)
    user = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
    if not user or not user.is_active or not _verify_password(payload.password, user.password_hash):
        _log_audit_event(db, None, "login_failed", "user", email, "Login failed", {"email": email})
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = _create_session_token(user)
    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        max_age=SESSION_TTL_SECONDS,
        httponly=True,
        secure=APP_ENV in {"production", "prod"},
        samesite="lax",
        path="/",
    )
    _log_audit_event(db, user, "login_success", "user", user.id, "Login successful")
    db.commit()
    return {"user": _serialize_user(user), "must_change_password": bool(getattr(user, "must_change_password", False))}


@app.post("/auth/logout")
def logout(request: Request, response: Response, db: Session = Depends(get_db)) -> dict:
    user = _get_user_from_request(request, db)
    if user:
        _log_audit_event(db, user, "logout", "user", user.id, "Logout", commit=True)
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return {"ok": True}


@app.get("/auth/me")
def auth_me(user: User = Depends(get_current_user)) -> dict:
    return {"user": _serialize_user(user), "auth_required": REQUIRE_AUTH, "must_change_password": bool(getattr(user, "must_change_password", False))}


@app.post("/auth/change-password")
def change_password(payload: ChangePasswordPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    if not _verify_password(payload.current_password, user.password_hash):
        _log_audit_event(db, user, "password_change_failed", "user", user.id, "Password change failed")
        db.commit()
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    new_password = _validate_password_policy(payload.new_password, "new_password")
    if _verify_password(new_password, user.password_hash):
        raise HTTPException(status_code=400, detail="New password must be different from the current password")
    user.password_hash = _hash_password(new_password)
    user.must_change_password = False
    _log_audit_event(db, user, "password_changed", "user", user.id, "Password changed")
    db.commit()
    db.refresh(user)
    return {"user": _serialize_user(user)}

@app.on_event("startup")
def on_startup() -> None:
    logging.basicConfig(level=logging.INFO)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    Base.metadata.create_all(bind=engine)
    migrate_qa_reviews_table(engine)
    migrate_qa_coaching_actions_table(engine)
    migrate_pilot_feedback_tables(engine)
    migrate_calls_table(engine)
    migrate_stt_provider_configs_table(engine)
    migrate_telephony_ingestion_tables(engine)
    migrate_production_readiness_tables(engine)
    migrate_access_control_tables(engine)
    with SessionLocal() as db:
        _bootstrap_admin_user(db)
        db.execute(text("UPDATE users SET visibility_scope = 'all' WHERE role = 'admin' AND (visibility_scope IS NULL OR visibility_scope = '' OR visibility_scope = 'team')"))
        db.commit()
    if APP_ENV in {"production", "prod"} and REQUIRE_AUTH and not ADMIN_PASSWORD:
        logger.warning("ADMIN_PASSWORD is missing in production mode; set it before exposing CallQuanta.")
    logger.info("CallQuanta API started")


def _normalize_optional_text(value: str | None) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    return normalized or None


def _normalize_direction(value: str | None) -> str | None:
    normalized = _normalize_optional_text(value)
    if normalized is None:
        return None
    normalized = normalized.lower()
    if normalized not in {"inbound", "outbound", "internal", "unknown"}:
        raise HTTPException(status_code=400, detail="direction must be one of: inbound, outbound, internal, unknown")
    return normalized


def _original_filename_basename(filename: str | None) -> str:
    basename = re.split(r"[\\/]+", filename or "")[-1].strip()
    return "".join(char for char in basename if char.isprintable() and char not in {"\x7f"})


def _truncate_display_filename(filename: str, max_length: int = MAX_DISPLAY_FILENAME_LENGTH) -> str:
    if len(filename) <= max_length:
        return filename
    extension = Path(filename).suffix
    if extension and len(extension) < max_length:
        return f"{filename[: max_length - len(extension)]}{extension}"
    return filename[:max_length]


def _display_upload_filename(filename: str | None) -> str:
    display_name = _truncate_display_filename(_original_filename_basename(filename))
    if not display_name:
        raise HTTPException(status_code=400, detail="Invalid filename")
    return display_name


def _upload_extension(display_name: str) -> str:
    extension = Path(display_name).suffix.lower()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_UPLOAD_EXTENSIONS))
        raise HTTPException(status_code=400, detail=f"Unsupported file extension. Allowed extensions: {allowed}")
    return extension


def _safe_storage_filename(display_name: str, extension: str) -> str:
    safe_name = secure_filename(display_name)
    if not safe_name or Path(safe_name).suffix.lower() != extension:
        stem = Path(safe_name).stem if safe_name else ""
        safe_name = f"{stem or 'upload'}{extension}"
    return safe_name


def _validate_upload_file(file: UploadFile) -> tuple[str, str]:
    display_name = _display_upload_filename(file.filename)
    extension = _upload_extension(display_name)

    content_type = (file.content_type or "").split(";", 1)[0].strip().lower()
    if (
        content_type
        and content_type != "application/octet-stream"
        and content_type not in ALLOWED_UPLOAD_CONTENT_TYPES
    ):
        raise HTTPException(status_code=400, detail="Unsupported file type")
    return display_name, _safe_storage_filename(display_name, extension)


def _format_upload_limit_message() -> str:
    return (
        "Upload too large. "
        f"Max per file: {MAX_UPLOAD_BYTES_PER_FILE / 1024 / 1024:.0f} MB. "
        f"Max bulk upload: {MAX_BULK_UPLOAD_BYTES / 1024 / 1024:.0f} MB."
    )


def _upload_size(file: UploadFile) -> int | None:
    size = getattr(file, "size", None)
    return int(size) if isinstance(size, int) and size >= 0 else None


def _validate_known_upload_sizes(files: list[UploadFile]) -> None:
    known_sizes = [_upload_size(file) for file in files]
    if MAX_UPLOAD_BYTES_PER_FILE:
        oversized = [size for size in known_sizes if size is not None and size > MAX_UPLOAD_BYTES_PER_FILE]
        if oversized:
            raise HTTPException(status_code=413, detail=_format_upload_limit_message())
    if MAX_BULK_UPLOAD_BYTES and all(size is not None for size in known_sizes):
        if sum(size or 0 for size in known_sizes) > MAX_BULK_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail=_format_upload_limit_message())


async def _store_upload_file(file: UploadFile, safe_name: str, bulk_state: dict[str, int] | None = None) -> tuple[str, Path, int]:
    stored_name = f"{uuid4().hex}_{safe_name}"
    stored_path = UPLOAD_DIR / stored_name
    size = 0

    with stored_path.open("wb") as handle:
        while True:
            chunk = await file.read(UPLOAD_CHUNK_SIZE)
            if not chunk:
                break
            size += len(chunk)
            if bulk_state is not None:
                bulk_state["size"] = bulk_state.get("size", 0) + len(chunk)
            if MAX_UPLOAD_BYTES_PER_FILE and size > MAX_UPLOAD_BYTES_PER_FILE:
                handle.close()
                stored_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=_format_upload_limit_message())
            if bulk_state is not None and MAX_BULK_UPLOAD_BYTES and bulk_state.get("size", 0) > MAX_BULK_UPLOAD_BYTES:
                handle.close()
                stored_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=_format_upload_limit_message())
            handle.write(chunk)

    if size == 0:
        stored_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    return stored_name, stored_path, size




def _hash_ingestion_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _new_ingestion_token() -> str:
    return secrets.token_urlsafe(32)


def _verify_ingestion_token(token: str | None, token_hash: str | None) -> bool:
    if not token or not token_hash:
        return False
    return hmac.compare_digest(_hash_ingestion_token(token), token_hash)


def _bearer_or_header_token(authorization: str | None, x_callquanta_token: str | None) -> str | None:
    if x_callquanta_token:
        return x_callquanta_token.strip()
    if authorization and authorization.lower().startswith("bearer "):
        return authorization.split(" ", 1)[1].strip()
    return None


def _safe_ingestion_message(message: str | None, limit: int = 500) -> str | None:
    if not message:
        return None
    cleaned = re.sub(r"(token|authorization|api[_-]?key)=?[^\s,;]+", r"\1=[redacted]", str(message), flags=re.I)
    return cleaned[:limit]


def _sanitized_ingestion_payload(payload: TelephonyImportPayload | None) -> dict | None:
    if payload is None:
        return None
    data = payload.model_dump(mode="json")
    if INGESTION_PAYLOAD_LOGGING:
        return {k: ("[redacted]" if k in {"customer_phone", "agent_phone"} else v) for k, v in data.items()}
    return {
        "external_call_id": data.get("external_call_id"),
        "filename": data.get("filename"),
        "direction": data.get("direction"),
        "language": data.get("language"),
        "duration_seconds": data.get("duration_seconds"),
        "auto_transcribe": data.get("auto_transcribe"),
        "auto_analyze": data.get("auto_analyze"),
    }


def _log_ingestion_event(db: Session, *, integration_id: int | None, source_provider: str, external_call_id: str | None, event_type: str, status: str, message: str | None = None, payload: TelephonyImportPayload | None = None, call_id: int | None = None) -> None:
    db.add(IngestionEvent(
        integration_id=integration_id,
        source_provider=source_provider,
        external_call_id=external_call_id,
        event_type=event_type,
        status=status,
        message=_safe_ingestion_message(message),
        payload_json=_sanitized_ingestion_payload(payload),
        call_id=call_id,
    ))


def _serialize_telephony_integration(integration: TelephonyIntegration, token: str | None = None) -> dict:
    return {
        "id": integration.id,
        "name": integration.name,
        "provider_type": integration.provider_type,
        "is_active": integration.is_active,
        "token_configured": bool(integration.ingestion_token_hash),
        "token": token,
        "auto_transcribe": integration.auto_transcribe,
        "auto_analyze": integration.auto_analyze,
        "default_agent_name": integration.default_agent_name,
        "default_team": integration.default_team,
        "default_campaign": integration.default_campaign,
        "default_direction": integration.default_direction,
        "default_language": integration.default_language,
        "webhook_path": f"/integrations/telephony/webhook/{integration.id}",
        "created_at": integration.created_at.isoformat() if integration.created_at else None,
        "updated_at": integration.updated_at.isoformat() if integration.updated_at else None,
    }


def _serialize_ingestion_event(event: IngestionEvent) -> dict:
    return {
        "id": event.id,
        "integration_id": event.integration_id,
        "source_provider": event.source_provider,
        "external_call_id": event.external_call_id,
        "event_type": event.event_type,
        "status": event.status,
        "message": event.message,
        "call_id": event.call_id,
        "created_at": event.created_at.isoformat() if event.created_at else None,
    }


def _validate_telephony_payload(payload: TelephonyImportPayload) -> None:
    if not _normalize_optional_text(payload.external_call_id):
        raise HTTPException(status_code=422, detail="external_call_id is required")
    if not _normalize_optional_text(payload.recording_url):
        raise HTTPException(status_code=422, detail="recording_url is required")
    if not re.match(r"^https?://", payload.recording_url.strip(), flags=re.I):
        raise HTTPException(status_code=400, detail="recording_url must be an http(s) URL")
    if payload.duration_seconds is not None and payload.duration_seconds < 0:
        raise HTTPException(status_code=400, detail="duration_seconds must be non-negative")
    _normalize_direction(payload.direction)
    normalized_language = normalize_stt_language(payload.language)
    if normalized_language is not None and normalized_language not in SUPPORTED_STT_LANGUAGE_CODES:
        raise HTTPException(status_code=400, detail="Unsupported audio language")


def _apply_telephony_payload(call: Call, payload: TelephonyImportPayload, integration: TelephonyIntegration) -> None:
    call.agent_name = _normalize_optional_text(payload.agent_name) or integration.default_agent_name
    call.team = _normalize_optional_text(payload.team) or integration.default_team
    call.campaign = _normalize_optional_text(payload.campaign) or integration.default_campaign
    call.direction = _normalize_direction(payload.direction) or integration.default_direction
    call.language = normalize_stt_language(payload.language) or integration.default_language
    call.customer_phone = _normalize_optional_text(payload.customer_phone)
    call.agent_phone = _normalize_optional_text(payload.agent_phone)
    call.started_at = payload.started_at
    call.ended_at = payload.ended_at
    call.duration_seconds = payload.duration_seconds


def _apply_call_metadata(call: Call, metadata: CallMetadataPayload) -> None:
    call.agent_name = _normalize_optional_text(metadata.agent_name)
    call.team = _normalize_optional_text(metadata.team)
    call.campaign = _normalize_optional_text(metadata.campaign)
    call.direction = _normalize_direction(metadata.direction)
    normalized_language = normalize_stt_language(metadata.language)
    if normalized_language is not None and normalized_language not in SUPPORTED_STT_LANGUAGE_CODES:
        raise HTTPException(status_code=400, detail="Unsupported audio language")
    call.language = normalized_language


async def _create_uploaded_call(file: UploadFile, db: Session, metadata: CallMetadataPayload | None = None, bulk_state: dict[str, int] | None = None) -> Call:
    display_name, safe_name = _validate_upload_file(file)
    stored_name, stored_path, size = await _store_upload_file(file, safe_name, bulk_state)
    call = Call(
        filename=display_name,
        status="uploaded",
        stored_filename=stored_name,
        stored_path=str(stored_path),
        file_size_bytes=size,
        content_type=file.content_type,
    )
    if metadata:
        _apply_call_metadata(call, metadata)
    db.add(call)
    db.commit()
    db.refresh(call)
    return call


def _ensure_not_last_active_admin(db: Session, user: User) -> None:
    if user.role != "admin" or not user.is_active:
        return
    active_admin_count = db.execute(select(func.count(User.id)).where(User.role == "admin", User.is_active == True)).scalar_one()
    if int(active_admin_count) <= 1:
        raise HTTPException(status_code=400, detail="Cannot deactivate or demote the last active admin")


def _apply_user_payload(target: User, payload: UserAccessCreatePayload | UserAccessPatchPayload) -> dict:
    updates = payload.model_dump(exclude_unset=True)
    if "email" in updates and updates["email"] is not None:
        target.email = _validate_email(updates["email"])
    if "role" in updates and updates["role"] is not None:
        target.role = _normalize_role(updates["role"])
    if "display_name" in updates:
        target.display_name = _normalize_optional_text(updates.get("display_name"))
    if "team" in updates:
        target.team = _normalize_optional_text(updates.get("team"))
    if "agent_name" in updates:
        target.agent_name = _normalize_optional_text(updates.get("agent_name"))
    if "visibility_scope" in updates or not getattr(target, "visibility_scope", None):
        target.visibility_scope = _normalize_visibility_scope(updates.get("visibility_scope"), target.role)
    if "is_active" in updates and updates["is_active"] is not None:
        target.is_active = bool(updates["is_active"])
    if "must_change_password" in updates and updates["must_change_password"] is not None:
        target.must_change_password = bool(updates["must_change_password"])
    return updates


@app.get("/settings/users")
def list_users(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    users = db.execute(select(User).order_by(User.email.asc())).scalars().all()
    teams = db.execute(select(Call.team).where(Call.team.is_not(None), Call.team != "").distinct().order_by(Call.team.asc())).scalars().all()
    agents = db.execute(select(Call.agent_name).where(Call.agent_name.is_not(None), Call.agent_name != "").distinct().order_by(Call.agent_name.asc())).scalars().all()
    return {"items": [_serialize_user(item) for item in users], "roles": sorted(VALID_ROLES), "visibility_scopes": sorted(VALID_VISIBILITY_SCOPES), "teams": [t for t in teams if str(t).strip()], "agents": [a for a in agents if str(a).strip()]}


@app.post("/settings/users")
def create_user(payload: UserAccessCreatePayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    email = _validate_email(payload.email)
    if db.execute(select(User.id).where(User.email == email)).first():
        raise HTTPException(status_code=409, detail="User with this email already exists")
    role = _normalize_role(payload.role)
    password_was_generated = not bool(payload.password)
    temporary_password = _generate_temporary_password() if password_was_generated else _validate_password_policy(payload.password)
    target = User(email=email, password_hash=_hash_password(temporary_password), role=role, visibility_scope=_normalize_visibility_scope(payload.visibility_scope, role), is_active=payload.is_active, must_change_password=bool(payload.must_change_password or password_was_generated))
    target.display_name = _normalize_optional_text(payload.display_name)
    target.team = _normalize_optional_text(payload.team)
    target.agent_name = _normalize_optional_text(payload.agent_name)
    db.add(target)
    db.flush()
    _log_audit_event(db, user, "user_created", "user", target.id, f"Created user {target.email}", {"role": target.role, "visibility_scope": target.visibility_scope})
    db.commit()
    db.refresh(target)
    response = {"user": _serialize_user(target)}
    if password_was_generated:
        response["temporary_password"] = temporary_password
    return response


@app.patch("/settings/users/{user_id}")
def patch_user(user_id: int, payload: UserAccessPatchPayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.id == user.id and (payload.role is not None or payload.visibility_scope is not None):
        raise HTTPException(status_code=400, detail="Users cannot change their own role or visibility scope")
    if payload.email is not None:
        new_email = _validate_email(payload.email)
        existing_user = db.execute(select(User).where(User.email == new_email, User.id != target.id)).first()
        if existing_user:
            raise HTTPException(status_code=409, detail="User with this email already exists")
    if (payload.role is not None and _normalize_role(payload.role) != "admin") or payload.is_active is False:
        _ensure_not_last_active_admin(db, target)
    old_role = target.role
    updates = _apply_user_payload(target, payload)
    if old_role != target.role:
        _log_audit_event(db, user, "role_changed", "user", target.id, f"Changed role for {target.email}", {"old_role": old_role, "new_role": target.role})
    _log_audit_event(db, user, "user_updated", "user", target.id, f"Updated user {target.email}", {"fields": sorted(updates.keys())})
    db.commit()
    db.refresh(target)
    return {"user": _serialize_user(target)}


@app.post("/settings/users/{user_id}/reset-password")
def reset_user_password(user_id: int, payload: ResetPasswordPayload | None = None, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    temporary_password = (payload.temporary_password if payload else None) or _generate_temporary_password()
    temporary_password = _validate_password_policy(temporary_password, "temporary_password")
    target.password_hash = _hash_password(temporary_password)
    target.must_change_password = True
    _log_audit_event(db, user, "password_reset", "user", target.id, f"Reset password for {target.email}")
    db.commit()
    return {"user": _serialize_user(target), "temporary_password": temporary_password}


@app.patch("/settings/users/{user_id}/deactivate")
def deactivate_user(user_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    _ensure_not_last_active_admin(db, target)
    target.is_active = False
    _log_audit_event(db, user, "user_deactivated", "user", target.id, f"Deactivated user {target.email}")
    db.commit()
    db.refresh(target)
    return {"user": _serialize_user(target)}


@app.patch("/settings/users/{user_id}/activate")
def activate_user(user_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    target.is_active = True
    _log_audit_event(db, user, "user_activated", "user", target.id, f"Activated user {target.email}")
    db.commit()
    db.refresh(target)
    return {"user": _serialize_user(target)}


def serialize_audit_event(event: AuditEvent) -> dict:
    return {"id": event.id, "actor_user_id": event.actor_user_id, "actor_email": event.actor_email, "action": event.action, "entity_type": event.entity_type, "entity_id": event.entity_id, "message": event.message, "metadata_json": event.metadata_json, "created_at": event.created_at.isoformat() if event.created_at else None}


@app.get("/settings/audit-log")
def audit_log(action: str | None = None, actor_email: str | None = None, entity_type: str | None = None, created_from: str | None = None, created_to: str | None = None, limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    stmt = select(AuditEvent)
    if _clean_filter(action):
        stmt = stmt.where(AuditEvent.action == action.strip())
    if _clean_filter(actor_email):
        stmt = stmt.where(AuditEvent.actor_email.ilike(f"%{actor_email.strip()}%"))
    if _clean_filter(entity_type):
        stmt = stmt.where(AuditEvent.entity_type == entity_type.strip())
    from_dt = _parse_created_bound(created_from)
    to_dt = _parse_created_bound(created_to, end_of_day=True)
    if from_dt:
        stmt = stmt.where(AuditEvent.created_at >= from_dt)
    if to_dt:
        stmt = stmt.where(AuditEvent.created_at <= to_dt)
    events = db.execute(stmt.order_by(AuditEvent.created_at.desc(), AuditEvent.id.desc()).limit(limit)).scalars().all()
    return {"items": [serialize_audit_event(event) for event in events]}


@app.post("/calls/upload")
async def upload_call(
    file: UploadFile = File(...),
    agent_name: str | None = Form(None),
    team: str | None = Form(None),
    campaign: str | None = Form(None),
    direction: str | None = Form(None),
    language: str | None = Form(None),
    db: Session = Depends(get_db),
) -> dict:
    metadata = CallMetadataPayload(
        agent_name=agent_name,
        team=team,
        campaign=campaign,
        direction=direction,
        language=language,
    )
    _normalize_direction(metadata.direction)
    _validate_known_upload_sizes([file])
    call = await _create_uploaded_call(file, db, metadata)
    return {"id": call.id, "filename": call.filename, "status": call.status}


@app.get("/settings/upload-limits")
def upload_limits(user: User = Depends(get_current_user)) -> dict:
    return {
        "max_upload_bytes_per_file": MAX_UPLOAD_BYTES_PER_FILE,
        "max_bulk_upload_bytes": MAX_BULK_UPLOAD_BYTES,
        "allowed_extensions": sorted(ALLOWED_UPLOAD_EXTENSIONS),
    }




@app.get("/settings/telephony/integrations")
def list_telephony_integrations(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    integrations = db.execute(select(TelephonyIntegration).order_by(TelephonyIntegration.id.asc())).scalars().all()
    events = db.execute(select(IngestionEvent).order_by(IngestionEvent.created_at.desc(), IngestionEvent.id.desc()).limit(50)).scalars().all()
    return {
        "presets": TELEPHONY_PROVIDER_PRESETS,
        "saved": [_serialize_telephony_integration(item) for item in integrations],
        "events": [_serialize_ingestion_event(event) for event in events],
    }


@app.post("/settings/telephony/integrations")
def upsert_telephony_integration(payload: TelephonyIntegrationPayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    if payload.provider_type not in TELEPHONY_PROVIDER_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported provider_type")
    name = _normalize_optional_text(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    default_direction = _normalize_direction(payload.default_direction)
    default_language = normalize_stt_language(payload.default_language)
    if default_language is not None and default_language not in SUPPORTED_STT_LANGUAGE_CODES:
        raise HTTPException(status_code=400, detail="Unsupported default_language")

    integration = db.get(TelephonyIntegration, payload.id) if payload.id else None
    if integration is None:
        integration = TelephonyIntegration(name=name, provider_type=payload.provider_type)
        db.add(integration)
    token: str | None = None
    if payload.generate_token or not integration.ingestion_token_hash:
        token = _new_ingestion_token()
        integration.ingestion_token_hash = _hash_ingestion_token(token)
    integration.name = name
    integration.provider_type = payload.provider_type
    integration.is_active = payload.is_active
    integration.auto_transcribe = payload.auto_transcribe or payload.auto_analyze
    integration.auto_analyze = payload.auto_analyze
    integration.default_agent_name = _normalize_optional_text(payload.default_agent_name)
    integration.default_team = _normalize_optional_text(payload.default_team)
    integration.default_campaign = _normalize_optional_text(payload.default_campaign)
    integration.default_direction = default_direction
    integration.default_language = default_language
    _log_audit_event(db, user, "telephony_integration_saved", "telephony_integration", integration.id, f"Saved telephony integration {integration.name}", {"token_regenerated": bool(token)})
    db.commit()
    db.refresh(integration)
    return _serialize_telephony_integration(integration, token=token)


@app.post("/settings/telephony/integrations/{integration_id}/regenerate-token")
def regenerate_telephony_token(integration_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    integration = db.get(TelephonyIntegration, integration_id)
    if not integration:
        raise HTTPException(status_code=404, detail="Integration not found")
    token = _new_ingestion_token()
    integration.ingestion_token_hash = _hash_ingestion_token(token)
    db.commit()
    db.refresh(integration)
    return _serialize_telephony_integration(integration, token=token)


@app.post("/settings/telephony/integrations/{integration_id}/toggle")
def toggle_telephony_integration(integration_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    integration = db.get(TelephonyIntegration, integration_id)
    if not integration:
        raise HTTPException(status_code=404, detail="Integration not found")
    integration.is_active = not integration.is_active
    db.commit()
    db.refresh(integration)
    return _serialize_telephony_integration(integration)


@app.get("/integrations/telephony/events")
def list_ingestion_events(limit: int = Query(50, ge=1, le=200), db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    events = db.execute(select(IngestionEvent).order_by(IngestionEvent.created_at.desc(), IngestionEvent.id.desc()).limit(limit)).scalars().all()
    return {"events": [_serialize_ingestion_event(event) for event in events]}


def _ingest_telephony_payload(integration: TelephonyIntegration, payload: TelephonyImportPayload, db: Session) -> dict:
    _validate_telephony_payload(payload)
    external_call_id = _normalize_optional_text(payload.external_call_id) or ""
    source_provider = integration.provider_type or "generic_webhook"
    _log_ingestion_event(db, integration_id=integration.id, source_provider=source_provider, external_call_id=external_call_id, event_type="webhook_received", status="success", payload=payload)

    existing = db.execute(select(Call).where(Call.source_provider == source_provider, Call.external_call_id == external_call_id)).scalars().first()
    if existing:
        _log_ingestion_event(db, integration_id=integration.id, source_provider=source_provider, external_call_id=external_call_id, event_type="duplicate_ignored", status="success", call_id=existing.id)
        db.commit()
        return {"status": "duplicate", "call_id": existing.id, "external_call_id": external_call_id}

    display_name = _display_upload_filename(payload.filename or f"{external_call_id}.wav")
    auto_analyze = bool(payload.auto_analyze if payload.auto_analyze is not None else integration.auto_analyze)
    auto_transcribe = bool(payload.auto_transcribe if payload.auto_transcribe is not None else integration.auto_transcribe) or auto_analyze
    call = Call(
        filename=display_name,
        status="recording_download_pending",
        source="telephony",
        source_provider=source_provider,
        external_call_id=external_call_id,
        external_recording_url=payload.recording_url,
        ingestion_status="recording_download_pending",
        auto_analyze_after_transcription=auto_analyze,
    )
    _apply_telephony_payload(call, payload, integration)
    db.add(call)
    db.flush()
    queued, warning = _enqueue_job(RECORDING_QUEUE, {"call_id": call.id, "recording_url": payload.recording_url, "filename": payload.filename or display_name, "auto_transcribe": auto_transcribe, "integration_id": integration.id})
    _log_ingestion_event(db, integration_id=integration.id, source_provider=source_provider, external_call_id=external_call_id, event_type="recording_download_queued", status="success" if queued else "failed", message=warning, call_id=call.id)
    if not queued:
        call.status = "recording_download_failed"
        call.ingestion_status = "recording_download_failed"
        call.ingestion_error = warning
        call.last_error_type = "ingestion"
        call.last_error_message = warning
    db.commit()
    db.refresh(call)
    return {"status": "accepted", "call_id": call.id, "external_call_id": external_call_id, "ingestion_status": call.ingestion_status}


@app.post("/integrations/telephony/webhook/{integration_id}")
def telephony_webhook(integration_id: int, payload: TelephonyImportPayload, authorization: str | None = Header(None), x_callquanta_token: str | None = Header(None), db: Session = Depends(get_db)) -> dict:
    integration = db.get(TelephonyIntegration, integration_id)
    if not integration or not integration.is_active:
        raise HTTPException(status_code=404, detail="Integration not found")
    token = _bearer_or_header_token(authorization, x_callquanta_token)
    if not _verify_ingestion_token(token, integration.ingestion_token_hash):
        _log_ingestion_event(db, integration_id=integration.id, source_provider=integration.provider_type or "generic_webhook", external_call_id=getattr(payload, "external_call_id", None), event_type="webhook_received", status="failed", message="Invalid ingestion token")
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid ingestion token")
    return _ingest_telephony_payload(integration, payload, db)


@app.post("/calls/upload/bulk")
async def upload_calls_bulk(
    files: list[UploadFile] = File(...),
    agent_name: str | None = Form(None),
    team: str | None = Form(None),
    campaign: str | None = Form(None),
    direction: str | None = Form(None),
    language: str | None = Form(None),
    db: Session = Depends(get_db),
) -> dict:
    metadata = CallMetadataPayload(
        agent_name=agent_name,
        team=team,
        campaign=campaign,
        direction=direction,
        language=language,
    )
    # Validate metadata and known sizes before storing any files so hard failures are returned consistently.
    _normalize_direction(metadata.direction)
    _validate_known_upload_sizes(files)

    uploaded: list[dict] = []
    failed: list[dict] = []
    bulk_state = {"size": 0}
    for file in files:
        filename = _truncate_display_filename(_original_filename_basename(file.filename)) or "unknown"
        try:
            call = await _create_uploaded_call(file, db, metadata, bulk_state)
            uploaded.append({"id": call.id, "filename": call.filename, "status": call.status})
        except HTTPException as exc:
            db.rollback()
            failed.append({"filename": filename, "error": str(exc.detail)})
        except Exception as exc:
            db.rollback()
            logger.exception("bulk upload failed for %s", filename)
            failed.append({"filename": filename, "error": str(exc) or "Upload failed"})
    return {"uploaded": uploaded, "failed": failed}


def serialize_call(call: Call) -> dict:
    return {
        "id": call.id,
        "filename": call.filename,
        "status": call.status,
        "stored_filename": call.stored_filename,
        "stored_path": call.stored_path,
        "file_size_bytes": call.file_size_bytes,
        "content_type": call.content_type,
        "agent_name": call.agent_name,
        "team": call.team,
        "campaign": call.campaign,
        "direction": call.direction,
        "language": call.language,
        "stt_provider_name": call.stt_provider_name,
        "stt_provider_type": call.stt_provider_type,
        "stt_model": call.stt_model,
        "stt_language_used": call.stt_language_used,
        "detected_language": call.detected_language,
        "source": call.source,
        "source_provider": call.source_provider,
        "external_call_id": call.external_call_id,
        "external_recording_url": call.external_recording_url,
        "customer_phone": call.customer_phone,
        "agent_phone": call.agent_phone,
        "started_at": call.started_at.isoformat() if call.started_at else None,
        "ended_at": call.ended_at.isoformat() if call.ended_at else None,
        "duration_seconds": call.duration_seconds,
        "ingestion_status": call.ingestion_status,
        "ingestion_error": call.ingestion_error,
        "imported_at": call.imported_at.isoformat() if call.imported_at else None,
        "auto_analyze_after_transcription": call.auto_analyze_after_transcription,
        "audio_deleted": bool(call.audio_deleted),
        "audio_deleted_at": call.audio_deleted_at.isoformat() if call.audio_deleted_at else None,
        "created_at": call.created_at.isoformat() if call.created_at else None,
        "last_error_type": call.last_error_type,
        "last_error_message": call.last_error_message,
        "last_processed_at": call.last_processed_at.isoformat() if call.last_processed_at else None,
    }


_AUDIO_MIME_TYPES = {
    ".wav": "audio/wav",
    ".mp3": "audio/mpeg",
    ".m4a": "audio/mp4",
    ".mp4": "audio/mp4",
    ".aac": "audio/aac",
    ".ogg": "audio/ogg",
    ".opus": "audio/ogg",
    ".flac": "audio/flac",
    ".webm": "audio/webm",
}


def _audio_content_type(call: Call, path: Path) -> str:
    stored_type = (call.content_type or "").split(";", 1)[0].strip().lower()
    if stored_type in ALLOWED_UPLOAD_CONTENT_TYPES:
        return {
            "audio/x-wav": "audio/wav",
            "audio/mp3": "audio/mpeg",
            "application/ogg": "audio/ogg",
            "audio/opus": "audio/ogg",
            "video/webm": "audio/webm",
        }.get(stored_type, stored_type)
    return _AUDIO_MIME_TYPES.get(path.suffix.lower()) or mimetypes.guess_type(path.name)[0] or "application/octet-stream"


def _parse_byte_range(value: str, file_size: int) -> tuple[int, int]:
    match = re.fullmatch(r"bytes=(\d*)-(\d*)", value.strip())
    if not match or file_size <= 0:
        raise HTTPException(status_code=416, detail="Requested range is not satisfiable", headers={"Content-Range": f"bytes */{file_size}"})
    start_text, end_text = match.groups()
    if not start_text and not end_text:
        raise HTTPException(status_code=416, detail="Requested range is not satisfiable", headers={"Content-Range": f"bytes */{file_size}"})
    if not start_text:
        suffix_length = int(end_text)
        if suffix_length <= 0:
            raise HTTPException(status_code=416, detail="Requested range is not satisfiable", headers={"Content-Range": f"bytes */{file_size}"})
        start = max(file_size - suffix_length, 0)
        end = file_size - 1
    else:
        start = int(start_text)
        end = min(int(end_text), file_size - 1) if end_text else file_size - 1
    if start >= file_size or start > end:
        raise HTTPException(status_code=416, detail="Requested range is not satisfiable", headers={"Content-Range": f"bytes */{file_size}"})
    return start, end


def _stream_file_range(path: Path, start: int, length: int):
    with path.open("rb") as audio_file:
        audio_file.seek(start)
        remaining = length
        while remaining:
            chunk = audio_file.read(min(AUDIO_STREAM_CHUNK_SIZE, remaining))
            if not chunk:
                break
            remaining -= len(chunk)
            yield chunk


@app.get("/calls/{call_id}/audio")
def get_call_audio(
    call_id: int,
    request: Request,
    download: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)
    path = Path(call.stored_path) if call.stored_path else None
    if call.audio_deleted or not path or not path.is_file():
        raise HTTPException(status_code=404, detail="Audio file not found")

    file_size = path.stat().st_size
    content_type = _audio_content_type(call, path)
    headers = {"Accept-Ranges": "bytes"}
    status_code = 200
    start, end = 0, max(file_size - 1, 0)
    range_header = request.headers.get("range")
    if range_header:
        start, end = _parse_byte_range(range_header, file_size)
        status_code = 206
        headers["Content-Range"] = f"bytes {start}-{end}/{file_size}"
    content_length = end - start + 1 if file_size else 0
    headers["Content-Length"] = str(content_length)

    if download:
        original_name = Path(call.filename or path.name).name or path.name
        fallback_name = secure_filename(original_name) or "call-recording"
        headers["Content-Disposition"] = f"attachment; filename=\"{fallback_name}\"; filename*=UTF-8''{quote(original_name)}"
        _log_audit_event(db, user, "audio_downloaded", "call", call.id, f"Audio downloaded: {original_name}")
        db.commit()

    return StreamingResponse(
        _stream_file_range(path, start, content_length),
        status_code=status_code,
        media_type=content_type,
        headers=headers,
    )


@app.patch("/calls/{call_id}/metadata")
def patch_call_metadata(call_id: int, payload: CallMetadataPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_modify_call(call, user)

    _apply_call_metadata(call, payload)
    _log_audit_event(db, user, "metadata_edited", "call", call.id, "Call metadata edited")
    db.commit()
    db.refresh(call)
    return serialize_call(call)


def _group_key(value: str | None) -> str:
    return (value or "").strip() or "Unassigned"



CALL_FILTER_STATUSES = {
    "uploaded",
    "transcription_pending",
    "transcribing",
    "transcription_failed",
    "transcribed",
    "analysis_pending",
    "analyzing",
    "analysis_failed",
    "analyzed",
    "failed",
    "recording_download_pending",
    "recording_download_failed",
    "ingestion_failed",
}
CALL_SORT_COLUMNS = {
    "id": Call.id,
    "filename": Call.filename,
    "status": Call.status,
    "agent_name": Call.agent_name,
    "team": Call.team,
    "campaign": Call.campaign,
    "direction": Call.direction,
    "language": Call.language,
    "file_size_bytes": Call.file_size_bytes,
    "created_at": Call.created_at,
    "last_processed_at": Call.last_processed_at,
}


def _clean_filter(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned or cleaned.lower() == "all":
        return None
    return cleaned


def _parse_created_bound(value: str | None, end_of_day: bool = False) -> datetime | None:
    cleaned = _clean_filter(value)
    if not cleaned:
        return None
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", cleaned):
            parsed_date = datetime.fromisoformat(cleaned).date()
            return datetime.combine(parsed_date, datetime_time.max if end_of_day else datetime_time.min)
        return datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date filter: {value}")


def _apply_call_filters(
    stmt,
    *,
    q: str | None = None,
    status: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
):
    search = _clean_filter(q)
    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(or_(Call.filename.ilike(pattern), Call.agent_name.ilike(pattern), Call.team.ilike(pattern), Call.campaign.ilike(pattern)))

    exact_filters = [
        (Call.status, _clean_filter(status), False),
        (Call.agent_name, _clean_filter(agent_name), False),
        (Call.team, _clean_filter(team), False),
        (Call.campaign, _clean_filter(campaign), False),
        (Call.direction, _clean_filter(direction), True),
    ]
    for column, value, allow_empty_alias in exact_filters:
        if value is None:
            continue
        if allow_empty_alias and value.lower() in {"unknown", "empty", "auto"}:
            stmt = stmt.where(or_(column == value, column.is_(None), column == ""))
        else:
            stmt = stmt.where(column == value)

    language_value = _clean_filter(language)
    if language_value:
        normalized_language = normalize_language_code(language_value)
        lowered_language = func.lower(func.trim(Call.language))
        if normalized_language == "auto":
            stmt = stmt.where(or_(Call.language.is_(None), func.trim(Call.language) == "", func.trim(Call.language) == "-", lowered_language == "auto"))
        else:
            stmt = stmt.where(or_(lowered_language == normalized_language, lowered_language.like(f"{normalized_language}-%"), lowered_language.like(f"{normalized_language}_%")))

    from_dt = _parse_created_bound(created_from)
    to_dt = _parse_created_bound(created_to, end_of_day=True)
    if from_dt:
        stmt = stmt.where(Call.created_at >= from_dt)
    if to_dt:
        stmt = stmt.where(Call.created_at <= to_dt)
    return stmt


def _filtered_calls_statement(user: User | None = None, **filters):
    stmt = _apply_call_filters(select(Call), **filters)
    return apply_call_scope(stmt, user) if user is not None else stmt


def _call_filter_query_params(
    q: str | None,
    status: str | None,
    agent_name: str | None,
    team: str | None,
    campaign: str | None,
    direction: str | None,
    language: str | None,
    created_from: str | None,
    created_to: str | None,
) -> dict:
    return {
        "q": q,
        "status": status,
        "agent_name": agent_name,
        "team": team,
        "campaign": campaign,
        "direction": direction,
        "language": language,
        "created_from": created_from,
        "created_to": created_to,
    }


def _get_latest_successful_reviews(db: Session, user: User | None = None, **filters) -> tuple[dict[int, Call], list[QAReview], dict[int, QAReview]]:
    calls = db.execute(_filtered_calls_statement(user=user, **filters)).scalars().all()
    calls_by_id = {call.id: call for call in calls}
    if not calls_by_id:
        return calls_by_id, [], {}
    successful_reviews = db.execute(
        select(QAReview)
        .where(QAReview.call_id.in_(list(calls_by_id.keys())), QAReview.status == "success", QAReview.score.is_not(None))
        .order_by(QAReview.call_id.asc(), QAReview.created_at.desc(), QAReview.id.desc())
    ).scalars().all()
    latest_by_call: dict[int, QAReview] = {}
    for review in successful_reviews:
        latest_by_call.setdefault(review.call_id, review)
    return calls_by_id, successful_reviews, latest_by_call


@app.get("/dashboard/metrics")
def dashboard_metrics(
    q: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    filters = _call_filter_query_params(q, None, agent_name, team, campaign, direction, language, created_from, created_to)
    calls_by_id, successful_reviews, latest_by_call = _get_latest_successful_reviews(db, user=user, **filters)
    calls = list(calls_by_id.values())

    total_calls = len(calls)
    uploaded_calls = sum(1 for c in calls if c.status and c.status.startswith("uploaded"))
    transcribed_calls = sum(1 for c in calls if c.status and c.status.startswith("transcribed"))
    analyzed_calls = len(latest_by_call)
    analysis_failed_calls = sum(1 for c in calls if c.status == "analysis_failed")
    filtered_call_ids = set(calls_by_id)
    total_qa_reviews = db.execute(select(QAReview.id).where(QAReview.call_id.in_(filtered_call_ids))).all() if filtered_call_ids else []

    latest_scores = [float(r.score) for r in latest_by_call.values() if r.score is not None]
    average_score = round(sum(latest_scores) / len(latest_scores), 2) if latest_scores else None
    lowest_score = min(latest_scores) if latest_scores else None
    highest_score = max(latest_scores) if latest_scores else None

    latest_reviews = sorted(latest_by_call.values(), key=lambda r: ((r.created_at.isoformat() if r.created_at else ""), r.id), reverse=True)[:5]
    latest_reviews_payload = []
    for review in latest_reviews:
        call = calls_by_id.get(review.call_id)
        latest_reviews_payload.append({
            "review_id": review.id,
            "call_id": review.call_id,
            "filename": call.filename if call else None,
            "created_at": review.created_at.isoformat() if review.created_at else None,
            "score": review.score,
            "agent_name": _group_key(call.agent_name if call else None),
            "team": _group_key(call.team if call else None),
            "campaign": _group_key(call.campaign if call else None),
            "provider_name": review.provider_name,
            "model": review.model,
            "scorecard_name": review.scorecard_name,
            "report_language": review.report_language,
            "summary": review.summary,
        "human_summary": review.human_summary,
        "human_notes": review.human_notes,
        "human_reviewer_user_id": review.human_reviewer_user_id,
        "calibration_notes": review.calibration_notes,
        "coaching_actions": [_serialize_coaching_action(action) for action in db.execute(select(QACoachingAction).where(QACoachingAction.review_id == review.id).order_by(QACoachingAction.created_at.desc(), QACoachingAction.id.desc())).scalars().all()],
        })

    lowest_reviews = sorted(latest_by_call.values(), key=lambda r: (float(r.score), r.created_at, r.id))[:5]
    lowest_payload = []
    for review in lowest_reviews:
        call = calls_by_id.get(review.call_id)
        lowest_payload.append({
            "review_id": review.id,
            "call_id": review.call_id,
            "filename": call.filename if call else None,
            "created_at": review.created_at.isoformat() if review.created_at else None,
            "score": review.score,
            "agent_name": _group_key(call.agent_name if call else None),
            "team": _group_key(call.team if call else None),
            "campaign": _group_key(call.campaign if call else None),
            "summary": review.summary,
        })

    criteria_rollup: dict[str, dict] = {}
    for review in latest_by_call.values():
        for criterion in (review.criteria_breakdown or []):
            max_points = float(criterion.get("max_points") or 0)
            if max_points <= 0:
                continue
            title = str(criterion.get("title") or "Untitled criterion").strip() or "Untitled criterion"
            score = float(criterion.get("score") or 0)
            percent = (score / max_points) * 100 if max_points > 0 else 0
            severity = str(criterion.get("severity") or "").lower()
            item = criteria_rollup.setdefault(title, {"criterion_title": title, "reviews_count": 0, "score_sum": 0.0, "max_sum": 0.0, "percent_sum": 0.0, "warning_count": 0, "critical_count": 0})
            item["reviews_count"] += 1
            item["score_sum"] += score
            item["max_sum"] += max_points
            item["percent_sum"] += percent
            if severity == "warning":
                item["warning_count"] += 1
            if severity == "critical":
                item["critical_count"] += 1

    criteria_problem_summary = []
    for item in criteria_rollup.values():
        count = item["reviews_count"]
        criteria_problem_summary.append({
            "criterion_title": item["criterion_title"],
            "reviews_count": count,
            "average_score": round(item["score_sum"] / count, 2),
            "average_max_points": round(item["max_sum"] / count, 2),
            "average_percent": round(item["percent_sum"] / count, 2),
            "warning_count": item["warning_count"],
            "critical_count": item["critical_count"],
        })
    criteria_problem_summary.sort(key=lambda x: (x["average_percent"], -x["critical_count"], -x["warning_count"]))

    reviewed = [r for r in latest_by_call.values() if (r.review_status or "ai_generated") != "ai_generated"]
    approved = [r for r in latest_by_call.values() if r.review_status == "approved"]
    disputed = [r for r in latest_by_call.values() if r.review_status == "disputed"]
    needs_rework = [r for r in latest_by_call.values() if r.review_status == "needs_rework"]
    human_scores = [float(r.human_total_score) for r in latest_by_call.values() if r.human_total_score is not None]
    deltas = [abs(float(r.ai_human_score_delta)) for r in latest_by_call.values() if r.ai_human_score_delta is not None]
    disagreement_rollup: dict[str, dict] = {}
    for r in latest_by_call.values():
        for criterion in (r.criteria_breakdown or []):
            if criterion.get("human_agrees") is False:
                title = str(criterion.get("title") or "Untitled criterion").strip() or "Untitled criterion"
                item = disagreement_rollup.setdefault(title, {"criterion_title": title, "disagreements": 0})
                item["disagreements"] += 1
    top_disagreements = sorted(disagreement_rollup.values(), key=lambda item: item["disagreements"], reverse=True)[:10]
    calibration = {
        "ai_reviews_count": len(latest_by_call),
        "human_reviewed_count": len(reviewed),
        "approved_count": len(approved),
        "disputed_count": len(disputed),
        "reviews_needing_rework": len(needs_rework),
        "average_ai_score": average_score,
        "average_human_score": round(sum(human_scores) / len(human_scores), 2) if human_scores else None,
        "average_ai_human_delta": round(sum(deltas) / len(deltas), 2) if deltas else None,
        "calibration_samples_count": sum(1 for r in latest_by_call.values() if r.calibration_flag),
        "top_criteria_disagreement": top_disagreements,
    }

    def build_group(group_field: str):
        groups: dict[str, dict] = {}
        for c in calls:
            key = _group_key(getattr(c, group_field))
            groups.setdefault(key, {group_field: key, "calls_count": 0, "analyzed_calls_count": 0, "scores": [], "latest_review_at": None})
            groups[key]["calls_count"] += 1
        for call_id, review in latest_by_call.items():
            call = calls_by_id.get(call_id)
            if not call:
                continue
            key = _group_key(getattr(call, group_field))
            item = groups.setdefault(key, {group_field: key, "calls_count": 0, "analyzed_calls_count": 0, "scores": [], "latest_review_at": None})
            item["analyzed_calls_count"] += 1
            item["scores"].append(float(review.score))
            if not item["latest_review_at"] or (review.created_at and review.created_at > item["latest_review_at"]):
                item["latest_review_at"] = review.created_at
        rows = []
        for key, item in groups.items():
            scores = item["scores"]
            rows.append({
                group_field: key,
                "calls_count": item["calls_count"],
                "analyzed_calls_count": item["analyzed_calls_count"],
                "average_score": round(sum(scores) / len(scores), 2) if scores else None,
                "lowest_score": min(scores) if scores else None,
                "highest_score": max(scores) if scores else None,
                "latest_review_at": item["latest_review_at"].isoformat() if item["latest_review_at"] else None,
            })
        rows.sort(key=lambda x: str(x[group_field]).lower())
        return rows

    return {
        "summary": {
            "total_calls": total_calls,
            "uploaded_calls": uploaded_calls,
            "transcribed_calls": transcribed_calls,
            "analyzed_calls": analyzed_calls,
            "analysis_failed_calls": analysis_failed_calls,
            "total_qa_reviews": len(total_qa_reviews),
            "average_score": average_score,
            "lowest_score": lowest_score,
            "highest_score": highest_score,
        },
        "latest_reviews": latest_reviews_payload,
        "lowest_score_reviews": lowest_payload,
        "criteria_problem_summary": criteria_problem_summary,
        "qa_calibration": calibration,
        "agent_metrics": build_group("agent_name"),
        "team_metrics": build_group("team"),
        "campaign_metrics": build_group("campaign"),
    }


@app.get("/dashboard/agent-metrics")
def dashboard_agent_metrics(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    metrics = dashboard_metrics(db=db, user=user)
    return {"agents": metrics["agent_metrics"]}


@app.get("/calls")
def list_calls(
    q: str | None = None,
    status: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    filters = _call_filter_query_params(q, status, agent_name, team, campaign, direction, language, created_from, created_to)
    filtered = _filtered_calls_statement(user=user, **filters)
    total = db.execute(select(func.count()).select_from(filtered.subquery())).scalar_one()
    sort_column = CALL_SORT_COLUMNS.get(sort_by, Call.created_at)
    order = sort_column.asc() if sort_dir.lower() == "asc" else sort_column.desc()
    tie_breaker = Call.id.asc() if sort_dir.lower() == "asc" else Call.id.desc()
    calls = db.execute(filtered.order_by(order, tie_breaker).limit(limit).offset(offset)).scalars().all()
    return {"items": [serialize_call(call) for call in calls], "total": int(total), "limit": limit, "offset": offset}


@app.get("/calls/filter-options")
def call_filter_options(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    def distinct_values(column):
        rows = db.execute(apply_call_scope(select(column).where(column.is_not(None), column != ""), user).distinct().order_by(column.asc())).scalars().all()
        return [row for row in rows if str(row).strip()]

    statuses = set(CALL_FILTER_STATUSES)
    statuses.update(str(row) for row in db.execute(apply_call_scope(select(Call.status).where(Call.status.is_not(None)), user).distinct()).scalars().all() if str(row).strip())
    language_codes = {normalize_language_code(value) for value in distinct_values(Call.language)}
    language_codes.add("auto")
    language_catalog = {item["code"]: item for item in SUPPORTED_STT_LANGUAGES}
    language_order = [item["code"] for item in SUPPORTED_STT_LANGUAGES]
    languages = [
        {
            "code": code,
            "label": language_catalog.get(code, {}).get("label_en", code),
            "label_ru": language_catalog.get(code, {}).get("label_ru", code),
        }
        for code in language_order
        if code in language_codes
    ]
    return {
        "agents": distinct_values(Call.agent_name),
        "teams": distinct_values(Call.team),
        "campaigns": distinct_values(Call.campaign),
        "directions": distinct_values(Call.direction),
        "languages": languages,
        "statuses": sorted(statuses),
    }


def _calls_csv_rows(calls: list[Call]) -> list[list]:
    return [[
        call.id,
        call.filename,
        call.status,
        call.agent_name or "",
        call.team or "",
        call.campaign or "",
        call.direction or "",
        call.language or "",
        call.file_size_bytes or "",
        call.content_type or "",
        call.created_at.isoformat() if call.created_at else "",
        call.last_processed_at.isoformat() if call.last_processed_at else "",
        call.last_error_message or "",
        call.source_provider or "",
        call.external_call_id or "",
        call.customer_phone or "",
        call.agent_phone or "",
        call.started_at.isoformat() if call.started_at else "",
        call.ended_at.isoformat() if call.ended_at else "",
        call.duration_seconds or "",
        call.ingestion_status or "",
    ] for call in calls]


def _parse_call_ids(call_ids: str | None) -> list[int] | None:
    cleaned = _clean_filter(call_ids)
    if not cleaned:
        return None
    ids: list[int] = []
    for raw_id in cleaned.split(","):
        raw_id = raw_id.strip()
        if not raw_id:
            continue
        try:
            parsed = int(raw_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="call_ids must be a comma-separated list of integers")
        if parsed > 0:
            ids.append(parsed)
    return list(dict.fromkeys(ids))


def _filtered_calls_for_export(db: Session, filters: dict, call_ids: str | None = None, user: User | None = None) -> list[Call]:
    """Return export calls; explicit call_ids win over any filter parameters."""
    selected_ids = _parse_call_ids(call_ids)
    if selected_ids is not None:
        if not selected_ids:
            return []
        scoped_stmt = apply_call_scope(select(Call).where(Call.id.in_(selected_ids)), user) if user is not None else select(Call).where(Call.id.in_(selected_ids))
        calls = db.execute(scoped_stmt.order_by(Call.created_at.desc(), Call.id.desc())).scalars().all()
        visible_ids = {call.id for call in calls}
        forbidden_ids = [call_id for call_id in selected_ids if call_id not in visible_ids]
        if forbidden_ids:
            raise HTTPException(status_code=403, detail=f"Selected call IDs are outside your visibility scope: {forbidden_ids}")
        return calls
    return db.execute(_filtered_calls_statement(user=user, **filters).order_by(Call.created_at.desc(), Call.id.desc())).scalars().all()


@app.get("/calls/export")
def export_calls(
    format: str = Query("csv", pattern="^(xlsx|csv)$"),
    q: str | None = None,
    status: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    call_ids: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    calls = _filtered_calls_for_export(db, _call_filter_query_params(q, status, agent_name, team, campaign, direction, language, created_from, created_to), call_ids=call_ids, user=user)
    _log_audit_event(db, user, "export_triggered", "calls", None, "Calls export triggered", {"format": format, "count": len(calls)}, commit=True)
    headers = ["ID", "Filename", "Status", "Agent", "Team", "Campaign", "Direction", "Language", "File size bytes", "Content type", "Created at", "Last processed at", "Last error", "Source provider", "External call ID", "Customer phone", "Agent phone", "Started at", "Ended at", "Duration seconds", "Ingestion status"]
    safe_name = "callquanta-filtered-calls"
    if format == "csv":
        buf = StringIO(); buf.write("\ufeff")
        writer = csv.writer(buf); writer.writerow(headers); writer.writerows(_calls_csv_rows(calls))
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'})
    wb = Workbook(); ws = wb.active; ws.title = "Calls"; ws.append(headers)
    for row in _calls_csv_rows(calls): ws.append(row)
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'})



VALID_ASSIGNMENT_STATUSES = {"assigned", "in_review", "completed", "skipped"}
VALID_QUALITY_VALUES = {"good", "acceptable", "poor", "unusable"}
VALID_SCORE_AGREEMENTS = {"agree", "partially_agree", "disagree"}
VALID_SCORECARD_FITS = {"good", "needs_changes", "wrong_for_this_call_type"}
VALID_ISSUE_TAGS = {"stt_quality", "wrong_language", "speaker_separation", "qa_logic", "score_too_high", "score_too_low", "missing_evidence", "bad_scorecard_criterion", "ui_confusing", "export_problem", "other"}


def _latest_assignments(db: Session, review_ids: list[int]) -> dict[int, QAReviewAssignment]:
    if not review_ids:
        return {}
    rows = db.execute(select(QAReviewAssignment).where(QAReviewAssignment.review_id.in_(review_ids)).order_by(QAReviewAssignment.review_id.asc(), QAReviewAssignment.created_at.desc(), QAReviewAssignment.id.desc())).scalars().all()
    out = {}
    for row in rows:
        out.setdefault(row.review_id, row)
    return out


def _feedback_by_review(db: Session, review_ids: list[int]) -> dict[int, QAFeedback]:
    if not review_ids:
        return {}
    rows = db.execute(select(QAFeedback).where(QAFeedback.review_id.in_(review_ids)).order_by(QAFeedback.updated_at.desc(), QAFeedback.id.desc())).scalars().all()
    out = {}
    for row in rows:
        out.setdefault(row.review_id, row)
    return out


def _serialize_assignment(a: QAReviewAssignment | None, users: dict[int, User] | None = None) -> dict | None:
    if not a:
        return None
    user = (users or {}).get(a.assigned_to_user_id)
    return {"id": a.id, "review_id": a.review_id, "call_id": a.call_id, "assigned_to_user_id": a.assigned_to_user_id, "assigned_to_email": user.email if user else None, "assigned_by_user_id": a.assigned_by_user_id, "status": a.status, "due_date": a.due_date.isoformat() if a.due_date else None, "created_at": a.created_at.isoformat() if a.created_at else None, "updated_at": a.updated_at.isoformat() if a.updated_at else None}


def _feedback_status(f: QAFeedback | None) -> str:
    if not f:
        return "no_feedback"
    tags = f.issue_tags_json or []
    return "has_issue" if tags else "feedback_added"


def _serialize_feedback(f: QAFeedback | None) -> dict | None:
    if not f:
        return None
    return {"id": f.id, "review_id": f.review_id, "call_id": f.call_id, "created_by_user_id": f.created_by_user_id, "created_by_email": f.created_by_email, "transcript_quality": f.transcript_quality, "qa_analysis_quality": f.qa_analysis_quality, "score_agreement": f.score_agreement, "scorecard_fit": f.scorecard_fit, "ai_missed_something": bool(f.ai_missed_something), "ai_missed_comment": f.ai_missed_comment, "ai_false_positive": bool(f.ai_false_positive), "ai_false_positive_comment": f.ai_false_positive_comment, "useful_for_coaching": f.useful_for_coaching, "coaching_usefulness_comment": f.coaching_usefulness_comment, "overall_feedback": f.overall_feedback, "issue_tags": f.issue_tags_json or [], "feedback_status": _feedback_status(f), "created_at": f.created_at.isoformat() if f.created_at else None, "updated_at": f.updated_at.isoformat() if f.updated_at else None}


def _upsert_feedback(db: Session, review: QAReview, call: Call, payload: QAFeedbackPayload, user: User) -> QAFeedback:
    if user.role not in {"admin", "manager", "supervisor"}:
        raise HTTPException(status_code=403, detail="Only admins, managers, and supervisors can edit QA feedback")
    require_can_modify_call(call, user)
    f = db.execute(select(QAFeedback).where(QAFeedback.review_id == review.id, QAFeedback.created_by_user_id == user.id).order_by(QAFeedback.id.desc())).scalars().first()
    created = f is None
    if f is None:
        f = QAFeedback(review_id=review.id, call_id=call.id, created_by_user_id=user.id, created_by_email=user.email)
        db.add(f)
    data = payload.model_dump(exclude_unset=True)
    for field in ["transcript_quality", "qa_analysis_quality"]:
        if field in data and data[field] is not None and data[field] not in VALID_QUALITY_VALUES:
            raise HTTPException(status_code=400, detail=f"invalid {field}")
    if data.get("score_agreement") is not None and data["score_agreement"] not in VALID_SCORE_AGREEMENTS:
        raise HTTPException(status_code=400, detail="invalid score_agreement")
    if data.get("scorecard_fit") is not None and data["scorecard_fit"] not in VALID_SCORECARD_FITS:
        raise HTTPException(status_code=400, detail="invalid scorecard_fit")
    if "issue_tags" in data:
        tags = [str(t).strip() for t in (data.get("issue_tags") or []) if str(t).strip()]
        bad = [t for t in tags if t not in VALID_ISSUE_TAGS]
        if bad:
            raise HTTPException(status_code=400, detail=f"invalid issue_tags: {bad}")
        f.issue_tags_json = tags
    for field in ["transcript_quality", "qa_analysis_quality", "score_agreement", "scorecard_fit", "ai_missed_something", "ai_missed_comment", "ai_false_positive", "ai_false_positive_comment", "useful_for_coaching", "coaching_usefulness_comment", "overall_feedback"]:
        if field in data:
            setattr(f, field, data[field] if not isinstance(data[field], str) else _normalize_optional_text(data[field]))
    f.updated_at = _utcnow()
    _log_audit_event(db, user, "feedback_submitted" if created else "feedback_updated", "qa_feedback", f.id, "QA feedback saved", {"review_id": review.id, "call_id": call.id})
    db.commit(); db.refresh(f)
    return f

@app.get("/qa-reviews")
def list_qa_reviews(
    q: str | None = None,
    status: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    review_status: str | None = None,
    calibration_flag: bool | None = None,
    min_score: float | None = None,
    max_score: float | None = None,
    min_delta: float | None = None,
    assigned_to: str | None = None,
    feedback_status: str | None = None,
    view: str | None = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    filters = _call_filter_query_params(q, status, agent_name, team, campaign, direction, language, created_from, created_to)
    calls = db.execute(_filtered_calls_statement(user=user, **filters)).scalars().all()
    calls_by_id = {call.id: call for call in calls}
    if not calls_by_id:
        return {"items": [], "total": 0, "limit": limit, "offset": offset}
    stmt = select(QAReview).where(QAReview.call_id.in_(list(calls_by_id.keys())))
    if review_status:
        if review_status == "ai_generated":
            stmt = stmt.where(or_(QAReview.review_status == "ai_generated", QAReview.review_status.is_(None)))
        else:
            stmt = stmt.where(QAReview.review_status == review_status)
    if calibration_flag is not None:
        if calibration_flag is False:
            stmt = stmt.where(or_(QAReview.calibration_flag.is_(False), QAReview.calibration_flag.is_(None)))
        else:
            stmt = stmt.where(QAReview.calibration_flag.is_(True))
    if min_score is not None:
        stmt = stmt.where(QAReview.score >= min_score)
    if max_score is not None:
        stmt = stmt.where(QAReview.score <= max_score)
    if min_delta is not None:
        stmt = stmt.where(func.abs(QAReview.ai_human_score_delta) >= min_delta)
    if view == "my_assigned" or assigned_to == "me":
        stmt = stmt.join(QAReviewAssignment, QAReviewAssignment.review_id == QAReview.id).where(QAReviewAssignment.assigned_to_user_id == user.id, QAReviewAssignment.status.in_(["assigned", "in_review"]))
    elif assigned_to:
        try:
            assigned_user_id = int(assigned_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="assigned_to must be a user id or me")
        stmt = stmt.join(QAReviewAssignment, QAReviewAssignment.review_id == QAReview.id).where(QAReviewAssignment.assigned_to_user_id == assigned_user_id)
    if view == "disputed": stmt = stmt.where(QAReview.review_status == "disputed")
    if view == "needs_rework": stmt = stmt.where(QAReview.review_status == "needs_rework")
    if view == "calibration": stmt = stmt.where(QAReview.calibration_flag.is_(True))
    if view == "low_score": stmt = stmt.where(QAReview.score <= 70)
    if view == "ai_unreviewed": stmt = stmt.where(or_(QAReview.review_status == "ai_generated", QAReview.review_status.is_(None)))
    if feedback_status in {"feedback_added", "has_issue", "no_feedback"}:
        feedback_ids = [row[0] for row in db.execute(select(QAFeedback.review_id)).all()]
        issue_ids = [row[0] for row in db.execute(select(QAFeedback.review_id).where(QAFeedback.issue_tags_json.is_not(None))).all()]
        if feedback_status == "no_feedback": stmt = stmt.where(QAReview.id.not_in(feedback_ids or [-1]))
        elif feedback_status == "feedback_added": stmt = stmt.where(QAReview.id.in_(feedback_ids or [-1]))
        elif feedback_status == "has_issue": stmt = stmt.where(QAReview.id.in_(issue_ids or [-1]))
    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    reviews = db.execute(stmt.order_by(QAReview.created_at.desc(), QAReview.id.desc()).limit(limit).offset(offset)).scalars().all()
    review_ids = [r.id for r in reviews]
    assignments = _latest_assignments(db, review_ids)
    feedback_map = _feedback_by_review(db, review_ids)
    user_ids = {a.assigned_to_user_id for a in assignments.values()}
    users = {u.id: u for u in db.execute(select(User).where(User.id.in_(user_ids))).scalars().all()} if user_ids else {}
    items = []
    for review in reviews:
        call = calls_by_id.get(review.call_id)
        coaching_count, open_count = _coaching_counts(db, review.id)
        items.append({
            **serialize_review_compact(review),
            "call_id": review.call_id,
            "filename": call.filename if call else None,
            "agent_name": call.agent_name if call else None,
            "team": call.team if call else None,
            "campaign": call.campaign if call else None,
            "coaching_actions_count": coaching_count,
            "open_coaching_actions_count": open_count,
            "assignment": _serialize_assignment(assignments.get(review.id), users),
            "assigned_to_user_id": assignments.get(review.id).assigned_to_user_id if assignments.get(review.id) else None,
            "assigned_to_email": users.get(assignments.get(review.id).assigned_to_user_id).email if assignments.get(review.id) and users.get(assignments.get(review.id).assigned_to_user_id) else None,
            "feedback": _serialize_feedback(feedback_map.get(review.id)),
            "feedback_status": _feedback_status(feedback_map.get(review.id)),
        })
    return {"items": items, "total": int(total), "limit": limit, "offset": offset}

@app.get("/qa-reviews/export")
def export_filtered_reviews(
    format: str = Query("csv", pattern="^(xlsx|csv)$"),
    q: str | None = None,
    status: str | None = None,
    agent_name: str | None = None,
    team: str | None = None,
    campaign: str | None = None,
    direction: str | None = None,
    language: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    call_ids: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    calls = _filtered_calls_for_export(db, _call_filter_query_params(q, status, agent_name, team, campaign, direction, language, created_from, created_to), call_ids=call_ids, user=user)
    _log_audit_event(db, user, "export_triggered", "qa_reviews", None, "QA reviews export triggered", {"format": format, "call_count": len(calls)}, commit=True)
    calls_by_id = {call.id: call for call in calls}
    reviews = db.execute(select(QAReview).where(QAReview.call_id.in_(list(calls_by_id.keys()))).order_by(QAReview.created_at.desc(), QAReview.id.desc())).scalars().all() if calls_by_id else []
    headers = ["Review ID", "Call ID", "Filename", "Review created at", "Analysis status", "Review status", "AI score", "Human score", "AI-human delta", "Human reviewer", "Human reviewed at", "Agent", "Team", "Campaign", "Provider", "Model", "Scorecard", "AI summary", "Human summary", "Human notes", "Coaching actions", "Open coaching actions", "Calibration sample"]
    rows = []
    for r in reviews:
        coaching_count, open_count = _coaching_counts(db, r.id)
        call = calls_by_id[r.call_id]
        rows.append([r.id, r.call_id, call.filename, r.created_at.isoformat() if r.created_at else "", r.status, r.review_status or "ai_generated", r.score if r.score is not None else "", r.human_total_score if r.human_total_score is not None else "", r.ai_human_score_delta if r.ai_human_score_delta is not None else "", r.human_reviewer_email or "", r.human_reviewed_at.isoformat() if r.human_reviewed_at else "", call.agent_name or "", call.team or "", call.campaign or "", r.provider_name or "", r.model or "", r.scorecard_name or "", r.summary or "", r.human_summary or "", r.human_notes or "", coaching_count, open_count, bool(r.calibration_flag)])
    safe_name = "callquanta-filtered-qa-reviews"
    if format == "csv":
        buf = StringIO(); buf.write("\ufeff")
        writer = csv.writer(buf); writer.writerow(headers); writer.writerows(rows)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'})
    wb = Workbook(); ws = wb.active; ws.title = "QA reviews"; ws.append(headers)
    for row in rows: ws.append(row)
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'})


def _enqueue_job(queue_name: str, payload: dict) -> tuple[bool, str | None]:
    try:
        redis_client.lpush(queue_name, json.dumps(payload))
        return True, None
    except redis.RedisError as exc:
        logger.warning("Redis enqueue failed for %s: %s", queue_name, exc)
        return False, "Queue backend is unavailable"


def _queue_length(queue_name: str) -> tuple[int, str | None]:
    try:
        return int(redis_client.llen(queue_name)), None
    except redis.RedisError as exc:
        logger.warning("Redis LLEN failed for %s: %s", queue_name, exc)
        return 0, "Redis unavailable; queue lengths are reported as 0."


def _status_counts(db: Session, user: User | None = None) -> dict[str, int]:
    statuses = [
        "uploaded",
        "transcription_pending",
        "transcribing",
        "transcription_failed",
        "transcribed",
        "analysis_pending",
        "analyzing",
        "analysis_failed",
        "analyzed",
        "failed",
        "recording_download_pending",
        "recording_download_failed",
        "ingestion_failed",
    ]
    counts = {status: 0 for status in statuses}
    stmt = select(Call.status, func.count(Call.id))
    if user is not None:
        stmt = apply_call_scope(stmt, user)
    rows = db.execute(stmt.group_by(Call.status)).all()
    for status, count in rows:
        status_key = str(status)
        if status_key in counts:
            counts[status_key] = int(count)
    return counts


@app.get("/calls/processing-summary")
def calls_processing_summary(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    return _status_counts(db, user=user)


@app.get("/jobs/summary")
def jobs_summary(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    transcription_queue_length, transcription_warning = _queue_length(TRANSCRIPTION_QUEUE)
    qa_queue_length, qa_warning = _queue_length(QA_QUEUE)
    counts = _status_counts(db, user=user)
    warnings = [warning for warning in {transcription_warning, qa_warning} if warning]
    payload = {
        "transcription_queue_length": transcription_queue_length,
        "qa_queue_length": qa_queue_length,
        "processing": {
            "transcribing": counts["transcribing"],
            "analyzing": counts["analyzing"],
        },
        "failed": {
            "transcription_failed": counts["transcription_failed"],
            "analysis_failed": counts["analysis_failed"],
            "failed": counts["failed"],
        },
        "pending": {
            "transcription_pending": counts["transcription_pending"],
            "analysis_pending": counts["analysis_pending"],
        },
    }
    if warnings:
        payload["warning"] = " ".join(warnings)
    return payload


def _unique_call_ids(call_ids: list[int]) -> list[int]:
    seen: set[int] = set()
    unique: list[int] = []
    for call_id in call_ids:
        if call_id in seen:
            continue
        seen.add(call_id)
        unique.append(call_id)
    return unique


@app.post("/calls/batch/transcribe")
def batch_transcribe_calls(payload: BatchCallsPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    results: list[dict] = []
    for call_id in _unique_call_ids(payload.call_ids):
        call = db.get(Call, call_id)
        if not call:
            results.append({"call_id": call_id, "status": "not_found", "error": "Call not found"})
            continue
        try:
            require_can_modify_call(call, user)
        except HTTPException as exc:
            results.append({"call_id": call_id, "status": "forbidden", "error": str(exc.detail)})
            continue
        if call.status in {"transcription_pending", "transcribing"}:
            results.append({"call_id": call_id, "status": "skipped", "reason": "Transcription already pending or processing"})
            continue
        if call.status in {"analysis_pending", "analyzing"}:
            results.append({"call_id": call_id, "status": "skipped", "reason": "Call is already in another workflow"})
            continue
        call.status = "transcription_pending"
        call.last_error_type = None
        call.last_error_message = None
        db.commit()
        queued, warning = _enqueue_job(TRANSCRIPTION_QUEUE, {"call_id": call_id})
        if queued:
            results.append({"call_id": call_id, "status": "transcription_queued"})
        else:
            call.status = "transcription_failed"
            call.last_error_type = "queue_unavailable"
            call.last_error_message = warning
            db.commit()
            results.append({"call_id": call_id, "status": "failed", "reason": warning})
    return {"results": results}


@app.post("/calls/batch/analyze")
def batch_analyze_calls(payload: BatchCallsPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    results: list[dict] = []
    for call_id in _unique_call_ids(payload.call_ids):
        call = db.get(Call, call_id)
        if not call:
            results.append({"call_id": call_id, "status": "not_found", "error": "Call not found"})
            continue
        try:
            require_can_modify_call(call, user)
        except HTTPException as exc:
            results.append({"call_id": call_id, "status": "forbidden", "error": str(exc.detail)})
            continue
        if call.status in {"analysis_pending", "analyzing"}:
            results.append({"call_id": call_id, "status": "skipped", "reason": "Analysis already pending or processing"})
            continue
        has_segments = db.execute(select(TranscriptSegment.id).where(TranscriptSegment.call_id == call_id).limit(1)).first()
        if not has_segments:
            results.append({"call_id": call_id, "status": "skipped", "reason": "Call has no transcript segments"})
            continue
        call.status = "analysis_pending"
        call.last_error_type = None
        call.last_error_message = None
        db.commit()
        queued, warning = _enqueue_job(QA_QUEUE, {"call_id": call_id})
        if queued:
            results.append({"call_id": call_id, "status": "analysis_queued"})
        else:
            call.status = "analysis_failed"
            call.last_error_type = "queue_unavailable"
            call.last_error_message = warning
            db.commit()
            results.append({"call_id": call_id, "status": "failed", "reason": warning})
    return {"results": results}



@app.patch("/calls/batch/metadata")
def batch_patch_call_metadata(payload: BatchMetadataPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    updates = payload.model_dump(exclude={"call_ids"}, exclude_unset=True)
    updates = {key: value for key, value in updates.items() if value is not None and (not isinstance(value, str) or value.strip())}
    results: list[dict] = []
    for call_id in _unique_call_ids(payload.call_ids):
        call = db.get(Call, call_id)
        if not call:
            results.append({"call_id": call_id, "status": "not_found", "error": "Call not found"})
            continue
        try:
            require_can_modify_call(call, user)
        except HTTPException as exc:
            results.append({"call_id": call_id, "status": "forbidden", "error": str(exc.detail)})
            continue
        if "agent_name" in updates:
            call.agent_name = _normalize_optional_text(updates["agent_name"])
        if "team" in updates:
            call.team = _normalize_optional_text(updates["team"])
        if "campaign" in updates:
            call.campaign = _normalize_optional_text(updates["campaign"])
        if "direction" in updates:
            call.direction = _normalize_direction(updates["direction"])
        if "language" in updates:
            normalized_language = normalize_stt_language(updates["language"])
            if normalized_language is not None and normalized_language not in SUPPORTED_STT_LANGUAGE_CODES:
                raise HTTPException(status_code=400, detail="Unsupported audio language")
            call.language = normalized_language
        results.append({"call_id": call_id, "status": "updated"})
    _log_audit_event(db, user, "metadata_edited", "call", None, "Batch metadata edited", {"count": len(results)})
    db.commit()
    return {"results": results}


@app.delete("/calls/batch")
def batch_delete_calls(payload: BatchDeletePayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    results: list[dict] = []
    for call_id in _unique_call_ids(payload.call_ids):
        call = db.get(Call, call_id)
        if not call:
            results.append({"call_id": call_id, "status": "not_found", "error": "Call not found"})
            continue
        file_error = None
        if payload.delete_files and call.stored_path:
            try:
                path = Path(call.stored_path)
                if not path.is_absolute():
                    path = Path.cwd() / path
                if path.exists() and path.is_file():
                    path.unlink()
            except OSError as exc:
                file_error = str(exc)
        db.execute(delete(TranscriptSegment).where(TranscriptSegment.call_id == call_id))
        db.execute(delete(QAReview).where(QAReview.call_id == call_id))
        db.delete(call)
        _log_audit_event(db, user, "call_deleted", "call", call_id, "Call deleted")
        results.append({"call_id": call_id, "status": "deleted", **({"file_error": file_error} if file_error else {})})
    _log_audit_event(db, user, "batch_delete", "call", None, "Batch delete completed", {"count": len(results)})
    db.commit()
    return {"results": results}


@app.post("/calls/batch/retry-failed")
def batch_retry_failed_calls(payload: BatchCallsPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    results: list[dict] = []
    for call_id in _unique_call_ids(payload.call_ids):
        call = db.get(Call, call_id)
        if not call:
            results.append({"call_id": call_id, "status": "not_found", "error": "Call not found"})
            continue
        try:
            require_can_modify_call(call, user)
        except HTTPException as exc:
            results.append({"call_id": call_id, "status": "forbidden", "error": str(exc.detail)})
            continue
        if call.status in {"transcription_failed", "failed"}:
            call.status = "transcription_pending"
            call.last_error_type = None
            call.last_error_message = None
            db.commit()
            queued, warning = _enqueue_job(TRANSCRIPTION_QUEUE, {"call_id": call_id})
            if queued:
                results.append({"call_id": call_id, "status": "transcription_queued"})
            else:
                call.status = "transcription_failed"
                call.last_error_type = "queue_unavailable"
                call.last_error_message = warning
                db.commit()
                results.append({"call_id": call_id, "status": "failed", "reason": warning})
            continue
        if call.status == "analysis_failed":
            has_segments = db.execute(select(TranscriptSegment.id).where(TranscriptSegment.call_id == call_id).limit(1)).first()
            if not has_segments:
                results.append({"call_id": call_id, "status": "skipped", "reason": "Call has no transcript segments"})
                continue
            call.status = "analysis_pending"
            call.last_error_type = None
            call.last_error_message = None
            db.commit()
            queued, warning = _enqueue_job(QA_QUEUE, {"call_id": call_id})
            if queued:
                results.append({"call_id": call_id, "status": "analysis_queued"})
            else:
                call.status = "analysis_failed"
                call.last_error_type = "queue_unavailable"
                call.last_error_message = warning
                db.commit()
                results.append({"call_id": call_id, "status": "failed", "reason": warning})
            continue
        results.append({"call_id": call_id, "status": "skipped", "reason": "Call is not in a failed status"})
    return {"results": results}

@app.get("/calls/{call_id}")
def get_call(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)
    return serialize_call(call)


@app.post("/calls/{call_id}/transcribe")
def transcribe_call(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_modify_call(call, user)

    call.status = "transcription_pending"
    call.last_error_type = None
    call.last_error_message = None
    db.commit()

    queued, warning = _enqueue_job(TRANSCRIPTION_QUEUE, {"call_id": call_id})
    if not queued:
        call.status = "transcription_failed"
        call.last_error_type = "queue_unavailable"
        call.last_error_message = warning
        db.commit()
        raise HTTPException(status_code=503, detail=warning)
    return {"call_id": call_id, "status": "transcription_queued"}


@app.get("/calls/{call_id}/transcript")
def get_call_transcript(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)

    segments = db.execute(
        select(TranscriptSegment)
        .where(TranscriptSegment.call_id == call_id)
        .order_by(TranscriptSegment.start_ms.asc(), TranscriptSegment.id.asc())
    ).scalars().all()

    return {
        "call_id": call_id,
        "status": call.status,
        "segments": [
            {
                "id": segment.id,
                "speaker": segment.speaker,
                "start_ms": segment.start_ms,
                "end_ms": segment.end_ms,
                "text": segment.text,
            }
            for segment in segments
        ],
    }


@app.post("/calls/{call_id}/analyze")
def analyze_call(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_modify_call(call, user)

    has_segments = db.execute(select(TranscriptSegment.id).where(TranscriptSegment.call_id == call_id).limit(1)).first()
    if not has_segments:
        raise HTTPException(status_code=400, detail="Call has no transcript segments")

    call.status = "analysis_pending"
    call.last_error_type = None
    call.last_error_message = None
    db.commit()

    queued, warning = _enqueue_job(QA_QUEUE, {"call_id": call_id})
    if not queued:
        call.status = "analysis_failed"
        call.last_error_type = "queue_unavailable"
        call.last_error_message = warning
        db.commit()
        raise HTTPException(status_code=503, detail=warning)
    return {"call_id": call_id, "status": "analysis_queued"}


@app.get("/calls/{call_id}/qa")
def get_call_qa(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)

    review = db.execute(
        select(QAReview)
        .where(QAReview.call_id == call_id, QAReview.status == "success")
        .order_by(QAReview.created_at.desc(), QAReview.id.desc())
    ).scalars().first()
    if not review:
        return {"call_id": call_id, "status": call.status, "review": None}

    return {"call_id": call_id, "status": call.status, "review": serialize_review_full(review, db)}



VALID_REVIEW_STATUSES = {"ai_generated", "human_reviewed", "approved", "disputed", "needs_rework"}
VALID_COACHING_STATUSES = {"open", "in_progress", "done", "dismissed", "acknowledged"}


def _normalize_review_status(status: str | None) -> str:
    normalized = (status or "human_reviewed").strip().lower()
    if normalized not in VALID_REVIEW_STATUSES or normalized == "ai_generated":
        raise HTTPException(status_code=400, detail="review_status must be human_reviewed, approved, disputed, or needs_rework")
    return normalized


def _serialize_coaching_action(action: QACoachingAction) -> dict:
    return {
        "id": action.id,
        "review_id": action.review_id,
        "call_id": action.call_id,
        "agent_name": action.agent_name,
        "assigned_to_user_id": action.assigned_to_user_id,
        "created_by_user_id": action.created_by_user_id,
        "title": action.title,
        "description": action.description,
        "status": action.status,
        "due_date": action.due_date.isoformat() if action.due_date else None,
        "created_at": action.created_at.isoformat() if action.created_at else None,
        "updated_at": action.updated_at.isoformat() if action.updated_at else None,
    }


def _coaching_counts(db: Session, review_id: int) -> tuple[int, int]:
    actions = db.execute(select(QACoachingAction).where(QACoachingAction.review_id == review_id)).scalars().all()
    return len(actions), sum(1 for action in actions if action.status in {"open", "in_progress"})


def _apply_criterion_human_reviews(review: QAReview, criterion_payloads: list[CriterionHumanReviewPayload] | None) -> None:
    if not criterion_payloads:
        return
    criteria = [dict(item) for item in (review.criteria_breakdown or [])]
    for payload in criterion_payloads:
        target_index = payload.criterion_index
        if target_index is None and payload.criterion_id:
            for index, criterion in enumerate(criteria):
                if str(criterion.get("id") or "") == str(payload.criterion_id):
                    target_index = index
                    break
        if target_index is None or target_index < 0 or target_index >= len(criteria):
            continue
        criterion = dict(criteria[target_index])
        if payload.human_score is not None:
            criterion["human_score"] = payload.human_score
        if payload.human_comment is not None:
            criterion["human_comment"] = _normalize_optional_text(payload.human_comment)
        if payload.human_severity is not None:
            criterion["human_severity"] = _normalize_optional_text(payload.human_severity)
        if payload.human_agrees is not None:
            criterion["human_agrees"] = bool(payload.human_agrees)
        criteria[target_index] = criterion
    review.criteria_breakdown = criteria


def serialize_review_compact(review: QAReview) -> dict:
    has_metadata = any([review.provider_name, review.model, review.scorecard_name, review.report_language])
    has_details = bool(review.criteria_breakdown) or bool(review.findings)
    legacy_review = not has_metadata and not has_details
    return {
        "id": review.id,
        "created_at": review.created_at.isoformat() if review.created_at else None,
        "status": review.status,
        "score": review.score,
        "provider_name": review.provider_name,
        "model": review.model,
        "scorecard_name": review.scorecard_name,
        "report_language": review.report_language,
        "analysis_mode": review.analysis_mode,
        "legacy_review": legacy_review,
        "review_status": review.review_status or "ai_generated",
        "human_reviewer_email": review.human_reviewer_email,
        "human_reviewed_at": review.human_reviewed_at.isoformat() if review.human_reviewed_at else None,
        "human_total_score": review.human_total_score,
        "ai_human_score_delta": review.ai_human_score_delta,
        "calibration_flag": bool(review.calibration_flag),
    }


def _legacy_findings_table_exists(db: Session) -> bool:
    cached = getattr(db.bind, "_qa_findings_exists", None)
    if cached is not None:
        return bool(cached)
    try:
        db.execute(text("SELECT 1 FROM qa_findings LIMIT 1"))
        setattr(db.bind, "_qa_findings_exists", True)
        return True
    except Exception:
        setattr(db.bind, "_qa_findings_exists", False)
        return False


def _legacy_review_details(review: QAReview, db: Session) -> tuple[list[dict], list[dict], dict]:
    criteria: list[dict] = []
    findings: list[dict] = []
    metadata: dict[str, str] = {}
    if not _legacy_findings_table_exists(db):
        return criteria, findings, metadata
    rows = db.execute(
        text("SELECT severity, evidence FROM qa_findings WHERE qa_review_id = :review_id ORDER BY id ASC"),
        {"review_id": review.id},
    ).mappings().all()
    for row in rows:
        severity = row.get("severity") or "info"
        evidence = row.get("evidence") or ""
        if evidence.startswith("[criterion:") and "]" in evidence:
            marker, _, remainder = evidence.partition("]")
            criteria.append(
                {
                    "id": f"legacy_{len(criteria) + 1}",
                    "title": marker[len("[criterion:") :].strip() or "Legacy criterion",
                    "score": "",
                    "max_points": "",
                    "comment": remainder.strip() or "Legacy criterion detail.",
                    "evidence": evidence,
                    "severity": severity,
                }
            )
            continue
        if evidence.lower().startswith("analysis mode:"):
            for chunk in evidence.split(";"):
                key, _, value = chunk.partition(":")
                if value:
                    metadata[key.strip().lower()] = value.strip()
        findings.append({"severity": severity, "evidence": evidence})
    return criteria, findings, metadata


def serialize_review_full(review: QAReview, db: Session) -> dict:
    criteria = review.criteria_breakdown or []
    findings = review.findings or []
    metadata_fallback: dict[str, str] = {}
    if not criteria and not findings:
        criteria, findings, metadata_fallback = _legacy_review_details(review, db)
    return {
        **serialize_review_compact(review),
        "summary": review.summary,
        "human_summary": review.human_summary,
        "human_notes": review.human_notes,
        "human_reviewer_user_id": review.human_reviewer_user_id,
        "calibration_notes": review.calibration_notes,
        "coaching_actions": [_serialize_coaching_action(action) for action in db.execute(select(QACoachingAction).where(QACoachingAction.review_id == review.id).order_by(QACoachingAction.created_at.desc(), QACoachingAction.id.desc())).scalars().all()],
        "assignment": _serialize_assignment(_latest_assignments(db, [review.id]).get(review.id), {u.id: u for u in db.execute(select(User).where(User.id.in_([a.assigned_to_user_id for a in _latest_assignments(db, [review.id]).values()]))).scalars().all()}),
        "feedback": _serialize_feedback(_feedback_by_review(db, [review.id]).get(review.id)),
        "feedback_status": _feedback_status(_feedback_by_review(db, [review.id]).get(review.id)),
        "provider_preset": review.provider_preset or metadata_fallback.get("preset"),
        "provider_name": review.provider_name or metadata_fallback.get("provider"),
        "model": review.model or metadata_fallback.get("model"),
        "scorecard_name": review.scorecard_name or metadata_fallback.get("scorecard"),
        "report_language": review.report_language or metadata_fallback.get("report language"),
        "analysis_mode": review.analysis_mode or metadata_fallback.get("analysis mode"),
        "criteria": criteria,
        "findings": findings,
        "error_message": review.error_message,
    }



@app.patch("/calls/{call_id}/qa/reviews/{review_id}/human-review")
def save_human_review(call_id: int, review_id: int, payload: HumanReviewPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id:
        raise HTTPException(status_code=404, detail="Review not found")
    require_can_modify_call(call, user)
    old_status = review.review_status or "ai_generated"
    old_score = review.human_total_score
    new_status = _normalize_review_status(payload.review_status)
    review.review_status = new_status
    review.human_reviewer_user_id = user.id
    review.human_reviewer_email = user.email
    review.human_reviewed_at = _utcnow()
    review.human_total_score = payload.human_total_score
    review.human_summary = _normalize_optional_text(payload.human_summary)
    review.human_notes = _normalize_optional_text(payload.human_notes)
    if review.score is not None and review.human_total_score is not None:
        review.ai_human_score_delta = round(float(review.human_total_score) - float(review.score), 2)
    else:
        review.ai_human_score_delta = None
    if payload.calibration_flag is not None:
        review.calibration_flag = bool(payload.calibration_flag)
    if payload.calibration_notes is not None:
        review.calibration_notes = _normalize_optional_text(payload.calibration_notes)
    _apply_criterion_human_reviews(review, payload.criteria)
    _log_audit_event(db, user, "human_review_saved", "qa_review", review.id, "Human QA review saved", {"old_status": old_status, "new_status": new_status})
    if new_status == "approved" and old_status != "approved":
        _log_audit_event(db, user, "review_approved", "qa_review", review.id, "QA review approved")
    if new_status == "disputed" and old_status != "disputed":
        _log_audit_event(db, user, "review_disputed", "qa_review", review.id, "QA review disputed")
    if old_score != review.human_total_score:
        _log_audit_event(db, user, "score_overridden", "qa_review", review.id, "QA score overridden", {"ai_score": review.score, "human_score": review.human_total_score})
    if payload.calibration_flag is not None:
        _log_audit_event(db, user, "calibration_flag_changed", "qa_review", review.id, "Calibration flag changed", {"calibration_flag": review.calibration_flag})
    db.commit()
    db.refresh(review)
    return {"call_id": call_id, "review": serialize_review_full(review, db)}


@app.post("/calls/{call_id}/qa/reviews/{review_id}/coaching-actions")
def create_coaching_action(call_id: int, review_id: int, payload: CoachingActionPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id:
        raise HTTPException(status_code=404, detail="Review not found")
    require_can_modify_call(call, user)
    title = _normalize_optional_text(payload.title)
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    action = QACoachingAction(review_id=review.id, call_id=call.id, agent_name=call.agent_name, created_by_user_id=user.id, assigned_to_user_id=payload.assigned_to_user_id, title=title, description=_normalize_optional_text(payload.description), status="open", due_date=payload.due_date)
    db.add(action)
    db.flush()
    _log_audit_event(db, user, "coaching_action_created", "qa_coaching_action", action.id, "Coaching action created", {"review_id": review.id, "call_id": call.id})
    db.commit()
    db.refresh(action)
    return {"action": _serialize_coaching_action(action)}


@app.patch("/calls/{call_id}/qa/reviews/{review_id}/coaching-actions/{action_id}")
def patch_coaching_action(call_id: int, review_id: int, action_id: int, payload: CoachingActionPatchPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    review = db.get(QAReview, review_id)
    action = db.get(QACoachingAction, action_id)
    if not call or not review or not action or review.call_id != call_id or action.review_id != review_id or action.call_id != call_id:
        raise HTTPException(status_code=404, detail="Coaching action not found")
    if user.role == "agent":
        require_can_view_call(call, user)
        if payload.status not in {"acknowledged", "done"}:
            raise HTTPException(status_code=403, detail="Agents can only acknowledge or complete their own coaching actions")
    else:
        require_can_modify_call(call, user)
    old_status = action.status
    if payload.title is not None:
        title = _normalize_optional_text(payload.title)
        if not title:
            raise HTTPException(status_code=400, detail="title cannot be empty")
        action.title = title
    if payload.description is not None:
        action.description = _normalize_optional_text(payload.description)
    if payload.assigned_to_user_id is not None:
        action.assigned_to_user_id = payload.assigned_to_user_id
    if payload.due_date is not None:
        action.due_date = payload.due_date
    if payload.status is not None:
        status = payload.status.strip().lower()
        if status not in VALID_COACHING_STATUSES:
            raise HTTPException(status_code=400, detail="invalid coaching status")
        action.status = status
    action.updated_at = _utcnow()
    if old_status != action.status and action.status in {"done", "dismissed"}:
        _log_audit_event(db, user, "coaching_action_completed", "qa_coaching_action", action.id, "Coaching action closed", {"old_status": old_status, "new_status": action.status})
    else:
        _log_audit_event(db, user, "coaching_action_updated", "qa_coaching_action", action.id, "Coaching action updated", {"old_status": old_status, "new_status": action.status})
    db.commit()
    db.refresh(action)
    return {"action": _serialize_coaching_action(action)}


@app.post("/calls/{call_id}/qa/reviews/{review_id}/assignments")
def assign_qa_review(call_id: int, review_id: int, payload: QAReviewAssignmentPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id); review = db.get(QAReview, review_id); target = db.get(User, payload.assigned_to_user_id)
    if not call or not review or review.call_id != call_id: raise HTTPException(status_code=404, detail="Review not found")
    require_can_modify_call(call, user)
    if not target or not target.is_active: raise HTTPException(status_code=404, detail="Assigned user not found")
    if not _can_view_call(target, call): raise HTTPException(status_code=403, detail="Assigned review is outside user's scope")
    assignment = QAReviewAssignment(review_id=review.id, call_id=call.id, assigned_to_user_id=target.id, assigned_by_user_id=user.id, status="assigned", due_date=payload.due_date)
    db.add(assignment); db.flush()
    _log_audit_event(db, user, "review_assigned", "qa_review_assignment", assignment.id, "QA review assigned", {"review_id": review.id, "assigned_to_user_id": target.id})
    db.commit(); db.refresh(assignment)
    return {"assignment": _serialize_assignment(assignment, {target.id: target})}


@app.patch("/calls/{call_id}/qa/reviews/{review_id}/assignments/{assignment_id}")
def patch_qa_review_assignment(call_id: int, review_id: int, assignment_id: int, payload: QAReviewAssignmentPatchPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id); review = db.get(QAReview, review_id); assignment = db.get(QAReviewAssignment, assignment_id)
    if not call or not review or not assignment or review.call_id != call_id or assignment.review_id != review_id: raise HTTPException(status_code=404, detail="Assignment not found")
    if user.role == "admin" or assignment.assigned_to_user_id == user.id: require_can_view_call(call, user)
    else: require_can_modify_call(call, user)
    old = assignment.status
    if payload.status is not None:
        status = payload.status.strip().lower()
        if status not in VALID_ASSIGNMENT_STATUSES: raise HTTPException(status_code=400, detail="invalid assignment status")
        assignment.status = status
    if payload.due_date is not None: assignment.due_date = payload.due_date
    assignment.updated_at = _utcnow()
    if old != assignment.status and assignment.status == "completed": _log_audit_event(db, user, "assignment_completed", "qa_review_assignment", assignment.id, "QA assignment completed")
    else: _log_audit_event(db, user, "assignment_updated", "qa_review_assignment", assignment.id, "QA assignment updated")
    db.commit(); db.refresh(assignment)
    target = db.get(User, assignment.assigned_to_user_id)
    return {"assignment": _serialize_assignment(assignment, {target.id: target} if target else {})}


@app.get("/calls/{call_id}/qa/reviews/{review_id}/feedback")
def get_qa_feedback(call_id: int, review_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id); review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id: raise HTTPException(status_code=404, detail="Review not found")
    require_can_view_call(call, user)
    items = db.execute(select(QAFeedback).where(QAFeedback.review_id == review.id).order_by(QAFeedback.updated_at.desc(), QAFeedback.id.desc())).scalars().all()
    return {"items": [_serialize_feedback(f) for f in items], "feedback": _serialize_feedback(items[0] if items else None)}


@app.put("/calls/{call_id}/qa/reviews/{review_id}/feedback")
def put_qa_feedback(call_id: int, review_id: int, payload: QAFeedbackPayload, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id); review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id: raise HTTPException(status_code=404, detail="Review not found")
    feedback = _upsert_feedback(db, review, call, payload, user)
    return {"feedback": _serialize_feedback(feedback)}


@app.get("/dashboard/pilot-feedback")
def pilot_feedback_metrics(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    calls = db.execute(_filtered_calls_statement(user=user)).scalars().all(); calls_by_id = {c.id: c for c in calls}
    feedback = db.execute(select(QAFeedback).where(QAFeedback.call_id.in_(list(calls_by_id.keys())))).scalars().all() if calls_by_id else []
    review_ids = [f.review_id for f in feedback]
    reviews = {r.id: r for r in db.execute(select(QAReview).where(QAReview.id.in_(review_ids))).scalars().all()} if review_ids else {}
    def dist(field):
        out = {}
        for f in feedback:
            v = getattr(f, field) or "not_set"; out[v] = out.get(v, 0) + 1
        return out
    tags = {}
    for f in feedback:
        for tag in (f.issue_tags_json or []): tags[tag] = tags.get(tag, 0) + 1
    deltas = [abs(float(reviews[f.review_id].ai_human_score_delta)) for f in feedback if reviews.get(f.review_id) and reviews[f.review_id].ai_human_score_delta is not None]
    useful = [f.useful_for_coaching for f in feedback if f.useful_for_coaching is not None]
    return {"reviews_with_feedback": len({f.review_id for f in feedback}), "transcript_quality_distribution": dist("transcript_quality"), "qa_quality_distribution": dist("qa_analysis_quality"), "score_agreement_distribution": dist("score_agreement"), "useful_for_coaching_percent": round(100 * sum(1 for v in useful if v) / len(useful), 2) if useful else None, "top_issue_tags": sorted(([{"tag": k, "count": v} for k, v in tags.items()]), key=lambda x: x["count"], reverse=True)[:10], "reviews_with_stt_problems": sum(1 for f in feedback if "stt_quality" in (f.issue_tags_json or [])), "reviews_with_qa_logic_problems": sum(1 for f in feedback if "qa_logic" in (f.issue_tags_json or [])), "average_ai_human_delta_with_feedback": round(sum(deltas)/len(deltas),2) if deltas else None}


@app.get("/qa-feedback/export")
def export_qa_feedback(format: str = Query("csv", pattern="^(xlsx|csv)$"), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    calls = db.execute(_filtered_calls_statement(user=user)).scalars().all(); calls_by_id = {c.id: c for c in calls}
    feedback = db.execute(select(QAFeedback).where(QAFeedback.call_id.in_(list(calls_by_id.keys()))).order_by(QAFeedback.created_at.desc())).scalars().all() if calls_by_id else []
    reviews = {r.id: r for r in db.execute(select(QAReview).where(QAReview.id.in_([f.review_id for f in feedback]))).scalars().all()} if feedback else {}
    headers = ["review_id","call_id","agent","team","campaign","ai_score","human_score","transcript_quality","qa_analysis_quality","score_agreement","scorecard_fit","useful_for_coaching","issue_tags","comments","created_by","created_at"]
    rows=[]
    for f in feedback:
        c=calls_by_id.get(f.call_id); r=reviews.get(f.review_id); comments=" | ".join(x for x in [f.ai_missed_comment, f.ai_false_positive_comment, f.coaching_usefulness_comment, f.overall_feedback] if x)
        rows.append([f.review_id,f.call_id,c.agent_name if c else "",c.team if c else "",c.campaign if c else "",r.score if r and r.score is not None else "",r.human_total_score if r and r.human_total_score is not None else "",f.transcript_quality or "",f.qa_analysis_quality or "",f.score_agreement or "",f.scorecard_fit or "",f.useful_for_coaching if f.useful_for_coaching is not None else "", ",".join(f.issue_tags_json or []), comments, f.created_by_email or "", f.created_at.isoformat() if f.created_at else ""])
    _log_audit_event(db, user, "pilot_export_created", "qa_feedback", None, "QA feedback export created", {"format": format, "count": len(rows)}, commit=True)
    if format == "csv":
        buf=StringIO(); buf.write("\ufeff"); w=csv.writer(buf); w.writerow(headers); w.writerows(rows)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="callquanta-qa-feedback.csv"'})
    wb=Workbook(); ws=wb.active; ws.title="QA feedback"; ws.append(headers)
    for row in rows: ws.append(row)
    out=BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="callquanta-qa-feedback.xlsx"'})

@app.get("/calls/{call_id}/qa/reviews")
def list_call_reviews(call_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)
    reviews = db.execute(select(QAReview).where(QAReview.call_id == call_id).order_by(QAReview.created_at.desc(), QAReview.id.desc())).scalars().all()
    return {"call_id": call_id, "reviews": [serialize_review_compact(r) for r in reviews]}


def _sanitize_filename(name: str) -> str:
    return secure_filename(name) or "export"


def _review_rows(call: Call, reviews: list[QAReview]) -> list[dict]:
    rows = []
    for review in reviews:
        criteria = review.criteria_breakdown or []
        if not criteria:
            rows.append({"review": review, "criterion": {}})
            continue
        for idx, c in enumerate(criteria, start=1):
            rows.append({"review": review, "criterion": {**c, "index": idx}})
    return rows


@app.get("/calls/{call_id}/qa/reviews/export")
def export_call_reviews(call_id: int, format: str = Query("xlsx", pattern="^(xlsx|csv|json)$"), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    call = db.get(Call, call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    require_can_view_call(call, user)
    reviews = db.execute(select(QAReview).where(QAReview.call_id == call_id).order_by(QAReview.created_at.desc(), QAReview.id.desc())).scalars().all()
    return _export_reviews(call, reviews, format, f"callquanta-call-{call_id}-qa-history", db)


@app.get("/calls/{call_id}/qa/reviews/{review_id}")
def get_call_review(call_id: int, review_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict:
    call = db.get(Call, call_id)
    review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id:
        raise HTTPException(status_code=404, detail="Review not found")
    require_can_view_call(call, user)
    return {"call_id": call_id, "review": serialize_review_full(review, db)}


@app.get("/calls/{call_id}/qa/reviews/{review_id}/export")
def export_single_review(call_id: int, review_id: int, format: str = Query("xlsx", pattern="^(xlsx|csv|json)$"), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    call = db.get(Call, call_id)
    review = db.get(QAReview, review_id)
    if not call or not review or review.call_id != call_id:
        raise HTTPException(status_code=404, detail="Review not found")
    require_can_view_call(call, user)
    return _export_reviews(call, [review], format, f"callquanta-call-{call_id}-review-{review_id}", db)


def _export_reviews(call: Call, reviews: list[QAReview], format: str, filename_root: str, db: Session):
    safe_name = _sanitize_filename(filename_root)
    if format == "json":
        payload = [{"call_id": call.id, "filename": call.filename, **serialize_review_full(r, db)} for r in reviews]
        data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        return StreamingResponse(BytesIO(data), media_type="application/json", headers={"Content-Disposition": f'attachment; filename="{safe_name}.json"'})
    if format == "csv":
        buf = StringIO()
        buf.write("\ufeff")
        headers = [
            "Review ID", "Call ID", "Filename", "Created at", "Review status", "AI score", "Human score", "AI-human delta",
            "Provider", "Model", "Scorecard", "Report language",
            "Criterion #", "Criterion title", "Criterion AI score", "Criterion human score", "Human agrees",
            "Criterion max points", "Comment", "Evidence", "Human comment",
        ]
        writer = csv.writer(buf)
        writer.writerow(headers)

        for row in _review_rows(call, reviews):
            r = row["review"]
            c = row["criterion"]
            vals = [
                r.id,
                call.id,
                call.filename,
                r.created_at.isoformat() if r.created_at else "",
                r.review_status or "ai_generated",
                r.score if r.score is not None else "",
                r.human_total_score if r.human_total_score is not None else "",
                r.ai_human_score_delta if r.ai_human_score_delta is not None else "",
                r.provider_name or "",
                r.model or "",
                r.scorecard_name or "",
                r.report_language or "",
                c.get("index", ""),
                c.get("title", ""),
                c.get("score", ""),
                c.get("human_score", ""),
                c.get("human_agrees", ""),
                c.get("max_points", ""),
                c.get("comment", ""),
                c.get("evidence", ""),
                c.get("human_comment", ""),
            ]
            writer.writerow(vals)

        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )
    wb=Workbook(); ws=wb.active; ws.title="Summary"
    ws.append(["Review ID","Call ID","Filename","Created at","Status","Review status","AI score","Human score","AI-human delta","Human reviewer","Human reviewed at","Provider","Model","Scorecard","Report language","AI summary","Human summary","Human notes","Calibration sample","Calibration notes"])
    for r in reviews:
        ws.append([r.id,call.id,call.filename,r.created_at.isoformat() if r.created_at else "",r.status,r.review_status or "ai_generated",r.score,r.human_total_score,r.ai_human_score_delta,r.human_reviewer_email,r.human_reviewed_at.isoformat() if r.human_reviewed_at else "",r.provider_name,r.model,r.scorecard_name,r.report_language,r.summary,r.human_summary,r.human_notes,bool(r.calibration_flag),r.calibration_notes])
    wc=wb.create_sheet("Criteria breakdown"); wc.append(["Review ID","Criterion #","Criterion title","Severity / level","AI score","Human score","Human agrees","Human severity","Max points","Comment","Evidence","Human comment"])
    wf=wb.create_sheet("Findings"); wf.append(["Review ID","Severity / level","Finding text"])
    for r in reviews:
        for i,c in enumerate(r.criteria_breakdown or [], start=1):
            wc.append([r.id,i,c.get("title"),c.get("severity"),c.get("score"),c.get("human_score"),c.get("human_agrees"),c.get("human_severity"),c.get("max_points"),c.get("comment"),c.get("evidence"),c.get("human_comment")])
        for f in r.findings or []:
            wf.append([r.id,f.get("severity"),f.get("evidence")])
    out=BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'})






def _normalize_retention_settings(value: dict | None) -> dict:
    settings = dict(DEFAULT_RETENTION_SETTINGS)
    if isinstance(value, dict):
        settings.update({key: value.get(key) for key in settings if key in value})
    env_map = {
        "audio_days": "RETENTION_AUDIO_DAYS",
        "transcripts_days": "RETENTION_TRANSCRIPTS_DAYS",
        "qa_reviews_days": "RETENTION_QA_REVIEWS_DAYS",
        "ingestion_events_days": "RETENTION_INGESTION_EVENTS_DAYS",
    }
    for key, env_name in env_map.items():
        if settings.get(key) is None and os.environ.get(env_name):
            settings[key] = os.environ.get(env_name)
        if settings.get(key) in {"", "forever", "none"}:
            settings[key] = None
        if settings.get(key) is not None:
            try:
                settings[key] = int(settings[key])
            except (TypeError, ValueError):
                settings[key] = None
            if settings[key] is not None and settings[key] < 1:
                settings[key] = None
    return settings


def _get_retention_settings(db: Session) -> dict:
    setting = db.get(AppSetting, RETENTION_SETTINGS_KEY)
    return _normalize_retention_settings(setting.value if setting else None)


def _upsert_retention_settings(db: Session, payload: RetentionSettingsPayload) -> dict:
    settings = _normalize_retention_settings(payload.model_dump())
    setting = db.get(AppSetting, RETENTION_SETTINGS_KEY)
    if not setting:
        db.add(AppSetting(key=RETENTION_SETTINGS_KEY, value=settings))
    else:
        setting.value = settings
    db.commit()
    return settings


def _upload_dir_usage() -> dict:
    count = 0
    size = 0
    if UPLOAD_DIR.exists():
        for path in UPLOAD_DIR.rglob("*"):
            if path.is_file():
                count += 1
                try:
                    size += path.stat().st_size
                except OSError:
                    pass
    return {"files_count": count, "total_bytes": size}


def _worker_heartbeat_status(name: str) -> dict:
    key = f"worker:{name}:heartbeat"
    try:
        raw = redis_client.get(key)
    except redis.RedisError as exc:
        return {"name": name, "ok": False, "last_heartbeat": None, "age_seconds": None, "error": str(exc)[:200]}
    if not raw:
        return {"name": name, "ok": False, "last_heartbeat": None, "age_seconds": None, "warning": "No heartbeat recorded"}
    try:
        ts = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=UTC)
        age = int((_utcnow() - ts).total_seconds())
    except ValueError:
        return {"name": name, "ok": False, "last_heartbeat": raw, "age_seconds": None, "warning": "Invalid heartbeat timestamp"}
    stale_after = max(60, WORKER_HEARTBEAT_SECONDS * 2)
    return {"name": name, "ok": age <= stale_after, "last_heartbeat": ts.isoformat(), "age_seconds": age, "warning": "Heartbeat is stale" if age > stale_after else None}


def _system_status(db: Session) -> dict:
    db_ok, db_error = _check_db_ready()
    redis_ok, redis_error = _check_redis_ready()
    llm_providers = db.execute(select(ProviderConfig).order_by(ProviderConfig.id.asc())).scalars().all()
    active_llm = next((item for item in llm_providers if (item.config or {}).get("is_active")), None) or (llm_providers[0] if llm_providers else None)
    active_stt = db.execute(select(SttProviderConfig).where(SttProviderConfig.is_active == True).limit(1)).scalar_one_or_none()
    return {
        "api": {"status": "ok", "version": app.version, "app_env": APP_ENV, "require_auth": REQUIRE_AUTH},
        "postgres": {"ok": db_ok, "error": db_error},
        "redis": {"ok": redis_ok, "error": redis_error},
        "queues": {
            "transcription": {"name": TRANSCRIPTION_QUEUE, "length": _queue_length(TRANSCRIPTION_QUEUE)[0]},
            "qa": {"name": QA_QUEUE, "length": _queue_length(QA_QUEUE)[0]},
            "recording": {"name": RECORDING_QUEUE, "length": _queue_length(RECORDING_QUEUE)[0]},
        },
        "workers": {name: _worker_heartbeat_status(name) for name in ["stt-worker", "qa-worker", "recording-worker"]},
        "providers": {
            "llm": serialize_provider_config(active_llm) if active_llm else None,
            "stt": serialize_stt_provider_config(active_stt) if active_stt else None,
        },
        "upload_limits": upload_limits(),
        "storage": _upload_dir_usage(),
    }


def _retention_cutoff(days: int | None) -> datetime | None:
    return _utcnow() - timedelta(days=days) if days else None


def _retention_plan(db: Session, settings: dict) -> dict:
    plan = {"audio": {"count": 0, "bytes": 0}, "transcripts": {"count": 0}, "qa_reviews": {"count": 0}, "ingestion_events": {"count": 0}}
    audio_cutoff = _retention_cutoff(settings.get("audio_days"))
    if audio_cutoff:
        calls = db.execute(select(Call).where(Call.created_at < audio_cutoff, Call.audio_deleted == False, Call.stored_path.is_not(None))).scalars().all()
        for call in calls:
            plan["audio"]["count"] += 1
            try:
                plan["audio"]["bytes"] += Path(call.stored_path).stat().st_size if call.stored_path else 0
            except OSError:
                pass
    transcript_cutoff = _retention_cutoff(settings.get("transcripts_days"))
    if transcript_cutoff:
        plan["transcripts"]["count"] = db.execute(select(func.count(TranscriptSegment.id)).join(Call, TranscriptSegment.call_id == Call.id).where(Call.created_at < transcript_cutoff)).scalar() or 0
    qa_cutoff = _retention_cutoff(settings.get("qa_reviews_days"))
    if qa_cutoff:
        plan["qa_reviews"]["count"] = db.execute(select(func.count(QAReview.id)).where(QAReview.created_at < qa_cutoff)).scalar() or 0
    ingestion_cutoff = _retention_cutoff(settings.get("ingestion_events_days"))
    if ingestion_cutoff:
        plan["ingestion_events"]["count"] = db.execute(select(func.count(IngestionEvent.id)).where(IngestionEvent.created_at < ingestion_cutoff)).scalar() or 0
    return plan


@app.get("/system/status")
def system_status(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    return _system_status(db)


@app.get("/settings/retention")
def get_retention_settings(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    settings = _get_retention_settings(db)
    return {"settings": settings, "preview": _retention_plan(db, settings)}


@app.patch("/settings/retention")
def patch_retention_settings(payload: RetentionSettingsPayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    settings = _upsert_retention_settings(db, payload)
    _log_audit_event(db, user, "retention_settings_updated", "settings", "retention", "Retention settings updated", commit=True)
    return {"settings": settings, "preview": _retention_plan(db, settings)}


@app.post("/settings/retention/preview")
def preview_retention_cleanup(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    settings = _get_retention_settings(db)
    return {"settings": settings, "preview": _retention_plan(db, settings)}


@app.post("/settings/retention/run-cleanup")
def run_retention_cleanup(confirm: bool = Query(False), db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    if not confirm:
        raise HTTPException(status_code=400, detail="Cleanup requires confirm=true after reviewing the dry-run preview")
    settings = _get_retention_settings(db)
    plan = _retention_plan(db, settings)
    audio_cutoff = _retention_cutoff(settings.get("audio_days"))
    if audio_cutoff:
        calls = db.execute(select(Call).where(Call.created_at < audio_cutoff, Call.audio_deleted == False, Call.stored_path.is_not(None))).scalars().all()
        for call in calls:
            if call.stored_path:
                Path(call.stored_path).unlink(missing_ok=True)
            call.audio_deleted = True
            call.audio_deleted_at = _utcnow()
            call.stored_path = None
            call.stored_filename = None
            call.file_size_bytes = None
    transcript_cutoff = _retention_cutoff(settings.get("transcripts_days"))
    if transcript_cutoff:
        old_call_ids = select(Call.id).where(Call.created_at < transcript_cutoff)
        db.execute(delete(TranscriptSegment).where(TranscriptSegment.call_id.in_(old_call_ids)))
    qa_cutoff = _retention_cutoff(settings.get("qa_reviews_days"))
    if qa_cutoff:
        db.execute(delete(QAReview).where(QAReview.created_at < qa_cutoff))
    ingestion_cutoff = _retention_cutoff(settings.get("ingestion_events_days"))
    if ingestion_cutoff:
        db.execute(delete(IngestionEvent).where(IngestionEvent.created_at < ingestion_cutoff))
    _log_audit_event(db, user, "retention_cleanup_run", "settings", "retention", "Retention cleanup run", plan)
    db.commit()
    return {"ok": True, "deleted": plan}

@app.get("/settings/languages")
def list_languages() -> list[dict]:
    return LANGUAGE_CATALOG


@app.get("/settings/stt-languages")
def list_stt_languages() -> list[dict]:
    return SUPPORTED_STT_LANGUAGES



def serialize_stt_provider_config(provider: SttProviderConfig) -> dict:
    return {
        "id": provider.id,
        "name": provider.name,
        "provider_type": provider.provider_type,
        "preset": provider.preset,
        "base_url": provider.base_url or "",
        "model": provider.model or "",
        "timeout_seconds": provider.timeout_seconds,
        "is_active": bool(provider.is_active),
        "api_key_configured": bool(provider.api_key),
        "created_at": provider.created_at.isoformat() if provider.created_at else None,
        "updated_at": provider.updated_at.isoformat() if provider.updated_at else None,
    }


def _active_stt_provider(db: Session) -> SttProviderConfig | None:
    return db.execute(select(SttProviderConfig).where(SttProviderConfig.is_active.is_(True)).order_by(SttProviderConfig.id.asc())).scalars().first()


def _validate_stt_provider_payload(payload: SttProviderUpsertRequest) -> None:
    if payload.provider_type not in STT_PROVIDER_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported STT provider_type")
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="STT provider name is required")
    if not payload.model.strip() and payload.provider_type != "custom":
        raise HTTPException(status_code=400, detail="STT model is required")
    if payload.provider_type in {"openai_compatible_audio", "groq_whisper", "deepgram", "assemblyai", "google_speech", "azure_speech"} and not (payload.api_key or payload.id):
        raise HTTPException(status_code=400, detail="API key is required for this hosted STT provider")
    if payload.timeout_seconds < 10:
        raise HTTPException(status_code=400, detail="timeout_seconds must be at least 10")


def _stt_provider_test_response(provider_type: str, model: str, base_url: str | None, api_key: str | None, timeout_seconds: int) -> dict:
    started = time.perf_counter()
    latency_ms = lambda: int((time.perf_counter() - started) * 1000)
    provider_type = (provider_type or "").strip()
    if provider_type not in STT_PROVIDER_TYPES:
        return {"ok": False, "latency_ms": latency_ms(), "provider_error": "Unsupported STT provider_type"}
    if provider_type == "faster_whisper_local":
        return {
            "ok": True,
            "latency_ms": latency_ms(),
            "model": model or os.environ.get("FASTER_WHISPER_MODEL", "base"),
            "provider_type": provider_type,
            "note": f"Configuration is valid. Worker will load model on demand using device={os.environ.get('FASTER_WHISPER_DEVICE', 'cpu')}.",
        }
    if provider_type == "openai_compatible_audio":
        if not api_key:
            return {"ok": False, "latency_ms": latency_ms(), "model": model, "provider_error": "API key is required."}
        if not (base_url or "").strip():
            return {"ok": False, "latency_ms": latency_ms(), "model": model, "provider_error": "Base URL is required."}
        if not (model or "").strip():
            return {"ok": False, "latency_ms": latency_ms(), "model": model, "provider_error": "Model is required."}
        return {
            "ok": True,
            "latency_ms": latency_ms(),
            "model": model,
            "provider_type": provider_type,
            "note": "Configuration validated. Provider test does not upload audio; transcription will call /audio/transcriptions.",
        }
    if provider_type in {"groq_whisper", "deepgram", "assemblyai", "google_speech", "azure_speech"}:
        if not api_key:
            return {"ok": False, "latency_ms": latency_ms(), "model": model, "provider_error": "API key is required for this hosted STT preset."}
        return {
            "ok": True,
            "latency_ms": latency_ms(),
            "model": model,
            "provider_type": provider_type,
            "note": "Configuration saved for future use. This STT provider integration is not implemented yet.",
        }
    return {
        "ok": True,
        "latency_ms": latency_ms(),
        "model": model,
        "provider_type": provider_type,
        "note": "Custom STT provider settings validated as a placeholder. Runtime transcription is not implemented yet.",
    }


@app.get("/settings/stt/providers")
def list_stt_providers(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    providers = db.execute(select(SttProviderConfig).order_by(SttProviderConfig.id.asc())).scalars().all()
    return {"presets": STT_PROVIDER_PRESETS, "saved": [serialize_stt_provider_config(item) for item in providers]}


@app.post("/settings/stt/providers")
def upsert_stt_provider(payload: SttProviderUpsertRequest, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    _validate_stt_provider_payload(payload)
    provider = db.get(SttProviderConfig, payload.id) if payload.id else None
    if provider is None:
        provider = SttProviderConfig(name=payload.name.strip(), provider_type=payload.provider_type, preset=payload.preset, model=payload.model.strip(), timeout_seconds=payload.timeout_seconds)
        db.add(provider)
    api_key = payload.api_key if payload.api_key is not None else provider.api_key
    provider.name = payload.name.strip()
    provider.provider_type = payload.provider_type
    provider.preset = payload.preset.strip() or "custom"
    provider.model = payload.model.strip()
    provider.base_url = (payload.base_url or "").strip().rstrip("/") or None
    provider.api_key = api_key or None
    provider.timeout_seconds = int(payload.timeout_seconds)
    provider.is_active = bool(payload.is_active)
    if provider.is_active:
        db.flush()
        others = db.execute(select(SttProviderConfig).where(SttProviderConfig.id != provider.id)).scalars().all()
        for other in others:
            other.is_active = False
    _log_audit_event(db, user, "stt_provider_saved", "stt_provider", provider.id, f"Saved STT provider {provider.name}")
    if provider.is_active:
        _log_audit_event(db, user, "stt_provider_activated", "stt_provider", provider.id, f"Activated STT provider {provider.name}")
    db.commit()
    db.refresh(provider)
    return serialize_stt_provider_config(provider)


@app.post("/settings/stt/providers/{provider_id}/activate")
def activate_stt_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(SttProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="STT provider not found")
    providers = db.execute(select(SttProviderConfig)).scalars().all()
    for item in providers:
        item.is_active = item.id == provider_id
    _log_audit_event(db, user, "stt_provider_activated", "stt_provider", provider.id, f"Activated STT provider {provider.name}")
    db.commit()
    db.refresh(provider)
    return serialize_stt_provider_config(provider)


@app.post("/settings/stt/providers/test")
def test_stt_provider(payload: SttProviderTestRequest, user: User = Depends(require_admin)) -> dict:
    return _stt_provider_test_response(payload.provider_type, payload.model, payload.base_url, payload.api_key, payload.timeout_seconds)


@app.post("/settings/stt/providers/{provider_id}/test")
def test_saved_stt_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(SttProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="STT provider not found")
    return _stt_provider_test_response(provider.provider_type, provider.model, provider.base_url, provider.api_key, provider.timeout_seconds)


@app.delete("/settings/stt/providers/{provider_id}")
def delete_stt_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(SttProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="STT provider not found")
    db.delete(provider)
    db.commit()
    return {"ok": True}

@app.get("/settings/stt")
def get_stt_settings(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    active = _active_stt_provider(db)
    if active:
        return {
            "mode": active.provider_type,
            "model": active.model,
            "provider": serialize_stt_provider_config(active),
        }
    return {
        "mode": os.environ.get("STT_MODE", "placeholder"),
        "model": os.environ.get("FASTER_WHISPER_MODEL", "base"),
        "provider": None,
    }


@app.get("/settings/workspace")
def get_workspace_settings(db: Session = Depends(get_db)) -> dict:
    return _get_workspace_settings(db)


@app.patch("/settings/workspace")
def patch_workspace_settings(payload: WorkspaceSettingsPayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    return _upsert_workspace_settings(db, payload)

def serialize_provider_config(provider: ProviderConfig) -> dict:
    config = provider.config or {}
    return {
        "id": provider.id,
        "name": provider.name,
        "provider_type": provider.provider_type,
        "preset": config.get("preset", "custom"),
        "base_url": config.get("base_url", ""),
        "model": config.get("model", ""),
        "timeout_seconds": config.get("timeout_seconds", 180),
        "is_active": bool(config.get("is_active", False)),
        "api_key_configured": bool(config.get("api_key")),
    }


def _load_default_scorecard() -> dict:
    if not DEFAULT_SCORECARD_PATH.exists():
        message = f"Default scorecard file not found: {DEFAULT_SCORECARD_PATH}"
        logger.error(message)
        raise HTTPException(status_code=500, detail=message)

    try:
        with DEFAULT_SCORECARD_PATH.open("r", encoding="utf-8") as handle:
            data = yaml.safe_load(handle) or {}
    except OSError as exc:
        logger.exception("Failed to read default scorecard file: %s", DEFAULT_SCORECARD_PATH)
        raise HTTPException(status_code=500, detail=f"Failed to read default scorecard file: {DEFAULT_SCORECARD_PATH}") from exc

    if not isinstance(data, dict) or not isinstance(data.get("criteria"), list):
        raise HTTPException(status_code=500, detail="Default scorecard is invalid")
    data.setdefault("name", "Default Sales QA")
    data.setdefault("report_language", "workspace")
    return data


def _validate_scorecard(payload: ScorecardPayload) -> None:
    if not payload.criteria:
        raise HTTPException(status_code=400, detail="Criteria list must not be empty")
    total_max = 0.0
    for criterion in payload.criteria:
        if not criterion.title.strip():
            raise HTTPException(status_code=400, detail="Criterion title must not be empty")
        if criterion.max_points <= 0:
            raise HTTPException(status_code=400, detail="Criterion max_points must be greater than 0")
        total_max += criterion.max_points
    if total_max <= 0:
        raise HTTPException(status_code=400, detail="Total max score must be greater than 0")
    if payload.report_language in {"english", "russian", "same_as_transcript", "workspace"}:
        return
    if not payload.report_language.strip():
        raise HTTPException(status_code=400, detail="Invalid report_language")


@app.get("/settings/scorecard")
def get_scorecard_settings(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    scorecard = db.execute(select(ScorecardConfig).order_by(ScorecardConfig.id.desc())).scalars().first()
    if scorecard and isinstance(scorecard.config, dict):
        return scorecard.config
    return _load_default_scorecard()


@app.post("/settings/scorecard")
def upsert_scorecard_settings(payload: ScorecardPayload, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    _validate_scorecard(payload)
    scorecard = db.execute(select(ScorecardConfig).order_by(ScorecardConfig.id.desc())).scalars().first()
    if not scorecard:
        scorecard = ScorecardConfig(name=payload.name.strip() or "Scorecard", config={})
        db.add(scorecard)
    scorecard.name = payload.name.strip() or "Scorecard"
    scorecard.config = payload.model_dump()
    db.commit()
    db.refresh(scorecard)
    return scorecard.config


@app.post("/settings/scorecard/reset")
def reset_scorecard_settings(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    default_scorecard = _load_default_scorecard()
    payload = ScorecardPayload(**default_scorecard)
    _validate_scorecard(payload)

    scorecard = db.execute(select(ScorecardConfig).order_by(ScorecardConfig.id.desc())).scalars().first()
    if not scorecard:
        scorecard = ScorecardConfig(name=payload.name.strip() or "Scorecard", config={})
        db.add(scorecard)

    scorecard.name = payload.name.strip() or "Scorecard"
    scorecard.config = payload.model_dump()
    db.commit()
    db.refresh(scorecard)
    return scorecard.config


@app.get("/settings/llm/providers")
def list_llm_providers(db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    providers = db.execute(select(ProviderConfig).order_by(ProviderConfig.id.asc())).scalars().all()
    return {"presets": PROVIDER_PRESETS, "saved": [serialize_provider_config(item) for item in providers]}


@app.post("/settings/llm/providers")
def upsert_llm_provider(payload: ProviderUpsertRequest, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(ProviderConfig, payload.id) if payload.id else None
    if provider is None:
        provider = ProviderConfig(provider_type=payload.provider_type, name=payload.name, config={})
        db.add(provider)

    existing = provider.config or {}
    api_key = payload.api_key if payload.api_key is not None else existing.get("api_key", "")
    provider.provider_type = payload.provider_type
    provider.name = payload.name
    provider.config = {"preset": payload.preset, "base_url": payload.base_url.rstrip("/"), "model": payload.model, "api_key": api_key, "timeout_seconds": payload.timeout_seconds, "is_active": payload.is_active}

    if payload.is_active:
        others = db.execute(select(ProviderConfig).where(ProviderConfig.id != provider.id)).scalars().all()
        for other in others:
            other_config = other.config or {}
            other_config["is_active"] = False
            other.config = other_config

    _log_audit_event(db, user, "llm_provider_saved", "llm_provider", provider.id, f"Saved LLM provider {provider.name}")
    if payload.is_active:
        _log_audit_event(db, user, "llm_provider_activated", "llm_provider", provider.id, f"Activated LLM provider {provider.name}")
    db.commit()
    db.refresh(provider)
    return serialize_provider_config(provider)


@app.post("/settings/llm/providers/{provider_id}/activate")
def activate_llm_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(ProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="Provider not found")
    providers = db.execute(select(ProviderConfig)).scalars().all()
    for item in providers:
        config = item.config or {}
        config["is_active"] = item.id == provider_id
        item.config = config
    _log_audit_event(db, user, "llm_provider_activated", "llm_provider", provider.id, f"Activated LLM provider {provider.name}")
    db.commit()
    db.refresh(provider)
    return serialize_provider_config(provider)


@app.delete("/settings/llm/providers/{provider_id}")
def delete_llm_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(ProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="Provider not found")
    db.delete(provider)
    db.commit()
    return {"ok": True}


@app.post("/settings/llm/providers/test")
def test_llm_provider(payload: ProviderTestRequest, user: User = Depends(require_admin)) -> dict:
    return _provider_test_request(payload.provider_type, payload.base_url, payload.model, payload.api_key, payload.timeout_seconds)


@app.post("/settings/llm/providers/{provider_id}/test")
def test_saved_llm_provider(provider_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> dict:
    provider = db.get(ProviderConfig, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="Provider not found")
    config = provider.config or {}
    return _provider_test_request(
        provider.provider_type,
        str(config.get("base_url", "")),
        str(config.get("model", "")),
        config.get("api_key"),
        float(config.get("timeout_seconds", 60)),
    )
